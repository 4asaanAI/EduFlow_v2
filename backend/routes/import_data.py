from __future__ import annotations
from typing import Optional

import csv
import io
import uuid
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile

from database import get_db
from middleware.auth import get_current_user
from services.audit_service import write_audit_doc
from tenant import get_school_id

router = APIRouter(prefix="/api/import", tags=["import"])

MAX_IMPORT_BYTES = 5 * 1024 * 1024
REQUIRED_COLUMNS = ["name", "class", "section", "parent_name", "parent_phone"]
OPTIONAL_COLUMNS = ["date_of_birth", "address", "route_zone_id"]
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _import_user(request: Optional[Request] = None, user: Optional[dict] = None) -> dict:
    user = user or get_current_user(request)
    if user.get("role") == "owner" or (user.get("role") == "admin" and user.get("sub_category") == "it_tech"):
        return user
    raise HTTPException(403, "Forbidden")


def _file_extension(filename: str) -> str:
    lower = (filename or "").lower()
    for ext in SUPPORTED_EXTENSIONS:
        if lower.endswith(ext):
            return ext
    raise HTTPException(400, "Only CSV and XLSX files are supported")


async def _read_file(file: UploadFile) -> tuple[str, bytes]:
    ext = _file_extension(file.filename or "")
    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(413, "Import file must be 5MB or smaller")
    if not content:
        raise HTTPException(400, "Import file is empty")
    return ext, content


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(400, "CSV file must include a header row")
    field_map = {field: _normalize_header(field) for field in reader.fieldnames}
    return [
        {field_map[key]: _normalize_cell(value) for key, value in row.items()}
        for row in reader
    ]


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "XLSX import support is not installed")

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "XLSX file must include a header row")
    headers = [_normalize_header(cell) for cell in rows[0]]
    return [
        {headers[idx]: _normalize_cell(value) for idx, value in enumerate(row) if idx < len(headers)}
        for row in rows[1:]
    ]


def _parse_rows(ext: str, content: bytes) -> list[dict]:
    rows = _parse_csv(content) if ext == ".csv" else _parse_xlsx(content)
    return [row for row in rows if any(value for value in row.values())]


async def _validate_rows(db, rows: list[dict]) -> dict:
    errors = []
    valid_rows = []
    duplicates = []
    class_cache = {}

    for index, row in enumerate(rows, start=2):
        row_errors = []
        for column in REQUIRED_COLUMNS:
            if not row.get(column):
                row_errors.append({"row": index, "field": column, "message": "Required column is missing"})

        class_name = row.get("class", "")
        section = row.get("section", "")
        class_key = (class_name.lower(), section.lower())
        cls = class_cache.get(class_key)
        if class_key not in class_cache:
            cls = await db.classes.find_one({"name": class_name, "section": section}, {"_id": 0})
            class_cache[class_key] = cls
        if class_name and section and not cls:
            row_errors.append({"row": index, "field": "class", "message": "Class and section do not exist"})

        duplicate = None
        if cls and row.get("name"):
            duplicate = await db.students.find_one(
                {"name": row["name"], "class_id": cls["id"], "is_active": True},
                {"_id": 0},
            )

        if row_errors:
            errors.extend(row_errors)
            continue

        normalized = {
            "row": index,
            "name": row["name"],
            "class_id": cls["id"],
            "class": class_name,
            "section": section,
            "parent_name": row["parent_name"],
            "parent_phone": row["parent_phone"],
            "date_of_birth": row.get("date_of_birth") or None,
            "address": row.get("address") or None,
            "route_zone_id": row.get("route_zone_id") or None,
            "duplicate_student_id": duplicate.get("id") if duplicate else None,
        }
        if duplicate:
            duplicates.append(
                {
                    "row": index,
                    "student_id": duplicate.get("id"),
                    "name": row["name"],
                    "class": class_name,
                    "section": section,
                    "message": "Student already exists",
                }
            )
        valid_rows.append(normalized)

    return {
        "valid_rows": valid_rows,
        "valid_count": len(valid_rows),
        "error_count": len(errors),
        "errors": errors,
        "duplicates": duplicates,
        "duplicate_count": len(duplicates),
    }


def _student_doc(row: dict, user: dict) -> dict:
    now = datetime.now().isoformat()
    return {
        "id": str(uuid.uuid4()),
        "schoolId": get_school_id(),
        "class_id": row["class_id"],
        "name": row["name"],
        "admission_number": f"IMP{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:5].upper()}",
        "dob": row.get("date_of_birth"),
        "address": row.get("address"),
        "uses_transport": bool(row.get("route_zone_id")),
        "bus_route": row.get("route_zone_id"),
        "admission_date": datetime.now().strftime("%Y-%m-%d"),
        "status": "active",
        "is_active": True,
        "created_by": user.get("id"),
        "created_at": now,
        "updated_at": now,
    }


def _guardian_doc(student_id: str, row: dict) -> dict:
    now = datetime.now().isoformat()
    return {
        "id": str(uuid.uuid4()),
        "schoolId": get_school_id(),
        "student_id": student_id,
        "name": row["parent_name"],
        "relation": "Father",
        "phone": row["parent_phone"],
        "whatsapp_phone": row["parent_phone"],
        "is_primary": True,
        "created_at": now,
        "updated_at": now,
    }


def _guardian_update(row: dict) -> dict:
    now = datetime.now().isoformat()
    return {
        "schoolId": get_school_id(),
        "name": row["parent_name"],
        "relation": "Father",
        "phone": row["parent_phone"],
        "whatsapp_phone": row["parent_phone"],
        "is_primary": True,
        "updated_at": now,
    }


@router.post("/validate")
async def validate_import(request: Request = None, file: UploadFile = File(...), user: Optional[dict] = None):
    if hasattr(request, "filename") and not hasattr(file, "filename"):
        file, request = request, None
    user = _import_user(request, user)
    db = get_db()
    ext, content = await _read_file(file)
    rows = _parse_rows(ext, content)
    report = await _validate_rows(db, rows)
    return {key: value for key, value in report.items() if key != "valid_rows"}


@router.post("/commit")
async def commit_import(
    request: Request = None,
    file: UploadFile = File(...),
    overwrite_duplicates: bool = Form(False),
    user: Optional[dict] = None,
):
    if hasattr(request, "filename") and not hasattr(file, "filename"):
        file, request = request, None
    user = _import_user(request, user)
    db = get_db()
    ext, content = await _read_file(file)
    rows = _parse_rows(ext, content)
    report = await _validate_rows(db, rows)

    imported_count = 0
    skipped_count = report["error_count"]
    for row in report["valid_rows"]:
        duplicate_id = row.get("duplicate_student_id")
        if duplicate_id and not overwrite_duplicates:
            skipped_count += 1
            continue

        if duplicate_id:
            student_update = _student_doc(row, user)
            student_update["id"] = duplicate_id
            student_update.pop("created_at", None)
            student_update.pop("created_by", None)
            await db.students.update_one({"id": duplicate_id}, {"$set": student_update})
            guardian_update_doc = _guardian_update(row)
            # Try to update an existing Father or Parent guardian (from prior imports)
            existing_g = await db.guardians.find_one(
                {"student_id": duplicate_id, "is_primary": True, "relation": {"$in": ["Father", "Parent"]}}
            )
            if existing_g:
                await db.guardians.update_one(
                    {"id": existing_g["id"]},
                    {"$set": guardian_update_doc},
                )
            else:
                await db.guardians.update_one(
                    {"student_id": duplicate_id, "is_primary": True},
                    {
                        "$set": guardian_update_doc,
                        "$setOnInsert": {"id": str(uuid.uuid4()), "student_id": duplicate_id, "created_at": datetime.now().isoformat()},
                    },
                    upsert=True,
                )
        else:
            student = _student_doc(row, user)
            await db.students.insert_one({**student, "_id": student["id"]})
            guardian = _guardian_doc(student["id"], row)
            await db.guardians.insert_one({**guardian, "_id": guardian["id"]})
        imported_count += 1

    audit = {
        "entity_type": "student",
        "entity_id": "bulk_import",
        "action": "bulk_import",
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "file_name": file.filename,
        "triggered_by": user.get("id"),
        "timestamp": datetime.now().isoformat(),
    }
    await write_audit_doc(db, audit, school_id=get_school_id(), branch_id=user.get("branch_id"))

    return {
        "success": True,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "errors": report["errors"],
        "duplicates": report["duplicates"],
        "duplicate_count": report["duplicate_count"],
    }
