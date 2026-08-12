"""Spreadsheet import - the full-file path into the school's live records.

Why this exists (2026-08-08). Attaching a spreadsheet to the chat gave Flo at most
40,000 characters of it. On the school's own 1,878-row student export that was 63
children - 3.4% - so "is anything in this file missing from the database?" was answered
from a fragment, confidently and wrongly. The answer is NOT a bigger slice: a 1.1 MB
sheet does not belong in a conversation at any size, and a language model re-typing
1,878 rows will eventually mistype one. So the file goes to a parser instead.

This service reads **every row**, deterministically, and never through the model:

* `preview_import()` writes nothing and reports exactly what would change, row by row.
  It is what the confirm card is built from - a person approves real numbers.
* `apply_import()` performs that same plan. Same code, same matching, so what was
  approved is what happens.

Two rules protect the school's data, and both are deliberate:

1. **Fill gaps, never overwrite.** By default a value already in the database wins over
   the file. A spreadsheet is usually a stale export; silently replacing 1,876 phone
   numbers with older ones would be catastrophic and invisible. `overwrite=True` exists
   for when someone genuinely means it, and the confirm card says so in words.
2. **Match on admission number, never on name.** Two children share a name far more
   often than an admission number, and merging two pupils' records is not recoverable.
   Rows without a usable admission number are REPORTED, never guessed at.

Services raise domain exceptions, never `HTTPException`.
"""

from __future__ import annotations

import io
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from services.actor_context import ActorContext
from services.ai_action_policy import privileged_profile
from services.audit_service import write_audit_doc
from services.txn_context import session_kwargs as _txn_session_kwargs
from tenant import scoped_filter, scoped_query

logger = logging.getLogger(__name__)

# No row cap. The whole point of this service is that nothing is sampled or truncated;
# a limit here would reintroduce exactly the silent data loss it was written to remove.
# Memory is bounded by the 20 MB upload limit at the boundary instead.

SUPPORTED_SUFFIXES = (".xlsx", ".xls", ".csv")


class ImportValidationError(Exception):
    """Unusable file, unknown target, or nothing to map → HTTP 400."""


class ImportFileUnavailableError(Exception):
    """The stored file cannot be read back → HTTP 404."""


def _session_kwargs(session):
    return _txn_session_kwargs(session)


def _now() -> datetime:
    return datetime.now(timezone.utc)


# ─── Column mapping ──────────────────────────────────────────────────────────
# The school's export uses its previous vendor's headers. Mapping them here (rather
# than asking the model to guess per row) is what makes the import repeatable.

# ⚠️ THE TARGET NAME ON THE RIGHT MUST BE THE NAME THE RECORD ALREADY USES.
#
# Corrected 2026-08-11, and this was a live defect rather than tidying. Eleven entries
# here pointed at invented names for facts the student records already hold under the
# names the 2026-08-06 load gave them: `whatsapp_phone` beside the real `whatsapp`,
# `father_phone` beside `father_mobile`, `aadhaar_number` beside `aadhaar_no`, and so
# on. Importing the school's own export would have written a SECOND copy of each,
# leaving the record with two Aadhaar numbers and every screen reading the empty one.
#
# That is not hypothetical. It is exactly what made the first missing-data report claim
# all 1,842 children had no date of birth when 1,055 had one under a different name.
# Before adding a row here, confirm the target against
# `student_service.UPDATABLE_FIELDS`; a name that is not in that set is a new field and
# needs a deliberate decision, not a guess.
STUDENT_FIELD_MAP: Dict[str, str] = {
    "name": "name",
    "mobile": "phone",
    "whatsapp": "whatsapp",
    "alternatenumber": "alternate_number",
    "email": "email",
    "gender": "gender",
    "dob": "dob",
    "dateofbirth": "dob",
    "fathername": "father_name",
    "mothername": "mother_name",
    "fathermobile": "father_mobile",
    "mothermobile": "mother_mobile",
    "address": "address",
    "admissiondate": "admission_date",
    "nationality": "nationality",
    "religion": "religion",
    "category": "category",
    "caste": "caste",
    "bloodgroup": "blood_group",
    "aadharno": "aadhaar_no",
    "aadhaarno": "aadhaar_no",
    "penno": "pen_no",
    "apaarid": "apaar_id",
    "registrationno": "registration_no",
    # `SID` and `Username` were mapped here until 2026-08-11 and are now refused
    # outright (see DELIBERATELY_NOT_IMPORTED). They are the PREVIOUS vendor's
    # identifiers; `student_service` already records that editing them makes them a
    # lie, and an import is an edit. `Username` is the more dangerous of the two:
    # EduFlow has its own logins in `auth_users`, and a second "username" on a student
    # record would eventually be read as the one people sign in with.
    "transport": "bus_route",
    "house": "house",
    "previousschool": "attended_school",
    "motherqualification": "mother_qualification",
    "fatherqualification": "father_qualification",
    "motheroccupation": "mother_occupation",
    "fatheroccupation": "father_occupation",
    "bankname": "bank_name",
    "bankaccountno": "bank_account_no",
    "ifsccode": "bank_ifsc",

    # ── Added 2026-08-11 (Abhimanyu): the rest of the school's export ──────────
    # The 2026-08-06 load put 71 of these columns onto the records and they have been
    # editable one at a time ever since, but the IMPORT path never learned their
    # headers, so a re-export of the school's own file reported them as columns it did
    # not recognise. The record and the door into it now agree.
    "bankifsc": "bank_ifsc",
    "bankbranch": "bank_branch",
    "accountholder": "account_holder",
    "srno": "sr_no",
    "rollno": "roll_number",
    "tcno": "tc_no",
    "tcdate": "tc_date",
    "attendedschool": "attended_school",
    "attendedclass": "attended_class",
    "enrolledclass": "enrolled_class",
    "enrolledsession": "enrolled_session",
    "lastsession": "last_session",
    "schoolaffiliated": "school_affiliated",
    "dobapplicationno": "dob_application_no",
    "fatheraadharno": "father_aadhaar_no",
    "motheraadharno": "mother_aadhaar_no",
    "city": "city",
    "state": "state",
    "country": "country",
    "pincode": "pincode",
    "placeofbirth": "place_of_birth",
    "medium": "medium",
    "remark": "remark",
    "admissiontype": "admission_type",
    "type": "admission_type",
    "isbplstudent": "is_bpl_student",
    "hasdisability": "has_disability",
    "rteapplicationno": "rte_application_no",
    # Empty in every export so far, and kept anyway because the school intends to fill
    # them THROUGH this platform (Abhimanyu, 2026-08-11). A column that is blank today
    # is not the same as a column nobody wants.
    "stream": "stream",
    "height": "height_cm",
    "weight": "weight_kg",
    # Parent detail. The father's side carries data and the mother's does not, which is
    # a reason to keep both rather than half a pair: a school that can record one
    # parent's email and not the other's has a gap that looks like a rule.
    "fatherresidentialaddress": "father_residential_address",
    "motherresidentialaddress": "mother_residential_address",
    "fatherofficialaddress": "father_official_address",
    "motherofficialaddress": "mother_official_address",
    "fatheremail": "father_email",
    "motheremail": "mother_email",
    "enrolledyear": "enrolled_year",
    # Empty today. Kept because they are the same kind of thing as the birth-certificate
    # and Right to Education application numbers, which DO carry data, and because 646
    # of the school's children are recorded as OBC or SC, for whom these are what a
    # government scholarship is claimed against.
    "domicileapplicationno": "domicile_application_no",
    "incomeapplicationno": "income_application_no",
    "casteapplicationno": "caste_application_no",
}

# Headers seen in the school's export that are deliberately NOT imported, so that
# "unrecognised column" means something. Reported as out of scope rather than silently
# ignored, and each one has a reason (2026-08-11):
#
#   Fee columns (TransportFees, SchoolTotalFees, TotalDiscount, GrossTotalFees, Fine,
#     PaidFees, Discount, BalanceFees) - money is reconciled through the fee ledger and
#     `PROTECTED_FIELDS` below already refuses it. A spreadsheet must never be able to
#     change what a family owes.
#   IsRteStudent - decides whether a child owes ANY school fee. Same reasoning as the
#     fee columns: it is a government entitlement, set deliberately per child, never in
#     bulk from a stale sheet. It stays editable one record at a time.
#   Class, Section, Status, Dropout, DropoutDate, DropoutReason - a child's standing at
#     the school. Already protected; naming them here stops the next reader re-adding
#     them "for completeness".
#   Photo, FatherPhoto, MotherPhoto, GuardianPhoto - file paths inside the PREVIOUS
#     vendor's storage. They do not resolve here and would render as broken images.
#   Guardian* - EduFlow keeps guardians as their own records, one per guardian. Landing
#     them on the student would create a second, disagreeing copy.
#   CreatedAt, LastActive - when the record was made and last touched in the previous
#     system. `created_at` on an EduFlow record means something different, and the
#     2026-08-06 load had to rename this column to avoid destroying it.
#   ScholarshipId, ScholarshipPassword - a credential. Never.
#   SamagraId - a Madhya Pradesh scheme identifier on a Uttar Pradesh school's export.
#   GeneralRegistrationNo, EnrollmentNo, SrnNo, BiometricCode, Reference, PanNo,
#     GovtStudentId, GovtFamilyId, Official* bank fields, *Income - empty in every
#     export, no school process behind them, and no request for them. These are the
#     "random columns offered by the previous vendor" that must not be carried across
#     empty (Abhimanyu, 2026-08-11).
DELIBERATELY_NOT_IMPORTED = frozenset({
    "transportfees", "schooltotalfees", "totaldiscount", "grosstotalfees", "fine",
    "paidfees", "discount", "balancefees", "isrtestudent", "class", "section",
    "status", "dropout", "dropoutdate", "dropoutreason", "photo", "fatherphoto",
    "motherphoto", "guardianphoto", "guardianname", "guardianmobile", "guardianemail",
    "guardianaadharno", "guardianoccupation", "guardianqualification",
    "guardianincome", "guardianresidentialaddress", "guardianofficialaddress",
    "createdat", "lastactive", "sid", "username",
    "scholarshipid", "scholarshippassword", "samagraid",
    "generalregistrationno", "enrollmentno", "srnno", "biometriccode", "reference",
    "panno", "govtstudentid", "govtfamilyid", "officialbankname", "officialbankbranch",
    "officialbankaccountno", "officialbankifsc", "officialaccountholder", "officialupi",
    "motherincome", "fatherincome", "houseblock",
})

ADMISSION_HEADERS = ("admissionno", "admissionnumber", "admno")

# ─── Who may import which columns ────────────────────────────────────────────
# Owner and principal import the whole student record. The accountant and the
# management profile import only their own segment, so that widening import to
# them does not quietly widen it to every field on a child's record.
#
# The profile names come from `ai_action_policy.privileged_profile` on purpose:
# that function is already the single place deciding what these four profiles
# are, and a second copy of that mapping here would drift away from it.

# The accountant's segment: the money-adjacent identifiers, plus the contact
# numbers fee reminders are actually sent to (that profile already sends them).
FINANCE_IMPORT_FIELDS = frozenset({
    "bank_name", "bank_account_no", "bank_ifsc", "bank_branch", "account_holder",
    "phone", "whatsapp", "alternate_number", "father_mobile", "mother_mobile",
})

# Bank details are the accountant's alone; management gets everything else.
_FINANCE_ONLY_FIELDS = frozenset({
    "bank_name", "bank_account_no", "bank_ifsc", "bank_branch", "account_holder",
})
NON_FINANCE_IMPORT_FIELDS = frozenset(STUDENT_FIELD_MAP.values()) - _FINANCE_ONLY_FIELDS

# None means "no field restriction" - not "no access". Access itself is decided
# by the route gate and the tool registry, never here.
IMPORT_FIELD_SCOPES: Dict[str, "frozenset | None"] = {
    "leadership": None,
    "finance": FINANCE_IMPORT_FIELDS,
    "non_finance": NON_FINANCE_IMPORT_FIELDS,
}

# Never let a spreadsheet touch these through this path: money is reconciled through
# the fee ledger, and enrolment status changes a child's standing at the school.
PROTECTED_FIELDS = frozenset({
    "id", "_id", "schoolId", "branch_id", "class_id", "status", "is_active",
    "admission_number", "fees", "balance", "paid_fees", "discount", "password_hash",
})


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("", "none", "null", "n/a", "na", "-") else text


def parse_rows(data: bytes, filename: str) -> List[Dict[str, str]]:
    """Parse EVERY data row. No sampling, no cap."""
    suffix = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if suffix not in SUPPORTED_SUFFIXES:
        raise ImportValidationError(
            f"'{filename}' is not a spreadsheet I can import. Supported: "
            f"{', '.join(SUPPORTED_SUFFIXES)}."
        )
    try:
        if suffix == ".csv":
            import csv

            text = data.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            return [{_norm_header(k): _clean(v) for k, v in row.items()} for row in reader]

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [_norm_header(h) for h in next(rows_iter)]
        except StopIteration:
            return []
        out: List[Dict[str, str]] = []
        for raw in rows_iter:
            if not any(v is not None and str(v).strip() for v in raw):
                continue
            out.append({h: _clean(v) for h, v in zip(header, raw) if h})
        return out
    except ImportValidationError:
        raise
    except Exception as exc:
        raise ImportValidationError(f"Could not read '{filename}': {exc}") from exc


def _admission_of(row: Dict[str, str]) -> str:
    for key in ADMISSION_HEADERS:
        if row.get(key):
            return row[key]
    return ""


def _mapped_values(row: Dict[str, str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for header, field in STUDENT_FIELD_MAP.items():
        if field in PROTECTED_FIELDS:
            continue
        value = row.get(header)
        if value:
            out[field] = value
    return out


async def _load_file(db, actor_ctx: ActorContext, file_id: str) -> tuple:
    record = await db.chat_uploaded_files.find_one(
        scoped_filter({"id": file_id}, actor_ctx.school_id), {"_id": 0}
    )
    if not record:
        raise ImportFileUnavailableError(
            f"I can't find an uploaded file with id '{file_id}'. Attach the file again."
        )
    key = record.get("s3_key")
    if not key:
        raise ImportFileUnavailableError(
            f"'{record.get('filename')}' was not stored, so I cannot read all of it. "
            "Please attach it again."
        )
    from services import s3_storage

    try:
        client = s3_storage.get_s3_client()
        obj = client.get_object(Bucket=s3_storage.get_bucket_name(), Key=key)
        return obj["Body"].read(), record
    except Exception as exc:
        raise ImportFileUnavailableError(
            f"'{record.get('filename')}' could not be read back from storage: {exc}"
        ) from exc


def allowed_import_fields(actor_ctx: ActorContext) -> "frozenset | None":
    """Which student fields this actor may import. None means every mapped field.

    Raises for any profile that is not one of the four reviewed ones, so that a
    role slipping past a route gate cannot import by default.
    """
    profile = privileged_profile({"role": actor_ctx.role, "sub_category": actor_ctx.sub_category})
    if profile not in IMPORT_FIELD_SCOPES:
        raise ImportValidationError("Importing a file is not part of your access.")
    return IMPORT_FIELD_SCOPES[profile]


async def _build_plan(db, actor_ctx: ActorContext, rows: List[Dict[str, str]],
                      *, overwrite: bool) -> Dict[str, Any]:
    """Work out, for every row, exactly what would change. Writes nothing."""
    allowed = allowed_import_fields(actor_ctx)
    admissions = [a for a in (_admission_of(r) for r in rows) if a]
    existing: Dict[str, dict] = {}
    if admissions:
        docs = await db.students.find(
            scoped_query({"admission_number": {"$in": admissions}},
                         branch_id=actor_ctx.branch_id),
            {"_id": 0},
        ).to_list(len(admissions) + 10)
        existing = {str(d.get("admission_number", "")).strip(): d for d in docs}

    updates: List[Dict[str, Any]] = []
    field_totals: Dict[str, int] = {}
    skipped_fields: set = set()
    unmatched: List[Dict[str, str]] = []
    no_admission: List[Dict[str, str]] = []
    seen: set = set()
    duplicates: List[str] = []

    for row in rows:
        admission = _admission_of(row)
        if not admission:
            no_admission.append({"name": row.get("name", ""), "reason": "no admission number in the row"})
            continue
        if admission in seen:
            duplicates.append(admission)
            continue
        seen.add(admission)

        student = existing.get(admission)
        if not student:
            unmatched.append({"admission_number": admission, "name": row.get("name", "")})
            continue

        changes: Dict[str, str] = {}
        # R2-18, 2026-08-11: the value each field held BEFORE the import, so the import
        # can actually be undone. The audit row's own comment already claimed it
        # "carries the BEFORE values"; it did not, it carried only the new ones, so an
        # import was the one thing Lalit does in bulk that could never be put back.
        previous: Dict[str, object] = {}
        for field, value in _mapped_values(row).items():
            # Out-of-scope columns are recorded and reported, never written. Dropping
            # them silently would let someone believe a column had been imported.
            if allowed is not None and field not in allowed:
                skipped_fields.add(field)
                continue
            current = str(student.get(field, "") or "").strip()
            if current and not overwrite:
                continue
            if current == value:
                continue
            changes[field] = value
            previous[field] = student.get(field)
        if changes:
            updates.append({
                "student_id": student.get("id"),
                "admission_number": admission,
                "name": student.get("name", ""),
                "changes": changes,
                "previous": previous,
            })
            for field in changes:
                field_totals[field] = field_totals.get(field, 0) + 1

    return {
        "rows_read": len(rows),
        "students_to_update": len(updates),
        "fields_to_fill": sum(field_totals.values()),
        "by_field": dict(sorted(field_totals.items(), key=lambda kv: -kv[1])),
        "rows_without_admission_number": len(no_admission),
        "duplicate_admission_numbers": len(duplicates),
        "not_found_in_database": len(unmatched),
        "not_found_sample": unmatched[:10],
        "overwrite": overwrite,
        "columns_outside_your_access": sorted(skipped_fields),
        "_updates": updates,
    }


def _rows_or_raise(data: bytes, filename: str) -> List[Dict[str, str]]:
    rows = parse_rows(data, filename)
    if not rows:
        raise ImportValidationError(f"'{filename}' has no data rows.")
    return rows


async def preview_upload(db, actor_ctx: ActorContext, data: bytes, filename: str,
                         *, overwrite: bool = False) -> Dict[str, Any]:
    """Preview a file uploaded straight from a screen, rather than a chat attachment.

    The screen and the chat share `_build_plan`, so the segment scoping, the
    fill-blanks-only rule and the admission-number matching cannot differ between them.
    """
    plan = await _build_plan(db, actor_ctx, _rows_or_raise(data, filename), overwrite=overwrite)
    plan["filename"] = filename
    plan.pop("_updates", None)
    return plan


async def apply_upload(db, actor_ctx: ActorContext, data: bytes, filename: str,
                       *, overwrite: bool = False, session=None) -> Dict[str, Any]:
    """Apply a file uploaded straight from a screen. Same writes as the chat path."""
    rows = _rows_or_raise(data, filename)
    return await _apply_plan(db, actor_ctx, rows, filename, overwrite=overwrite, session=session)


async def preview_import(db, actor_ctx: ActorContext, params: dict) -> Dict[str, Any]:
    """What an import WOULD do, across every row. Performs no writes."""
    data, record = await _load_file(db, actor_ctx, params.get("file_id", ""))
    rows = _rows_or_raise(data, record.get("filename", ""))
    plan = await _build_plan(db, actor_ctx, rows, overwrite=bool(params.get("overwrite")))
    plan["filename"] = record.get("filename", "")
    plan["file_id"] = params.get("file_id", "")
    plan.pop("_updates", None)
    return plan


async def apply_import(db, actor_ctx: ActorContext, params: dict, *, session=None) -> Dict[str, Any]:
    """Apply the plan the preview described. One audit row per student changed."""
    data, record = await _load_file(db, actor_ctx, params.get("file_id", ""))
    rows = _rows_or_raise(data, record.get("filename", ""))
    return await _apply_plan(db, actor_ctx, rows, record.get("filename", ""),
                             overwrite=bool(params.get("overwrite")), session=session)


async def _apply_plan(db, actor_ctx: ActorContext, rows: List[Dict[str, str]], filename: str,
                      *, overwrite: bool, session=None) -> Dict[str, Any]:
    """The write itself. Shared by the chat-attachment path and the screen upload."""
    record = {"filename": filename}
    plan = await _build_plan(db, actor_ctx, rows, overwrite=overwrite)
    updates = plan.pop("_updates", [])
    if not updates:
        outside = plan.get("columns_outside_your_access") or []
        # Saying "already has this information" when the real reason is that every
        # usable column was outside this profile's segment would be a lie.
        reason = (
            "Nothing was imported: the columns in this file are outside your access ("
            + ", ".join(outside) + "). Someone with wider access needs to import these."
            if outside and not plan.get("fields_to_fill")
            else "Nothing to change - the database already has this information."
        )
        plan.update({"filename": record.get("filename", ""), "applied": 0, "message": reason})
        return plan

    batch_id = str(uuid.uuid4())
    applied = 0
    failed = 0
    for item in updates:
        try:
            await db.students.update_one(
                scoped_filter({"id": item["student_id"]}, actor_ctx.school_id),
                {"$set": {**item["changes"], "updated_at": actor_ctx.now_iso(),
                          "updated_by": actor_ctx.user_id}},
                **_session_kwargs(session),
            )
            # Per-student audit carrying the BEFORE values, so an import can be
            # unpicked later. A batch-level summary alone would not be reversible.
            await write_audit_doc(
                db,
                {
                    "_id": str(uuid.uuid4()),
                    "id": str(uuid.uuid4()),
                    "schoolId": actor_ctx.school_id,
                    "entity_type": "student",
                    "entity_id": item["student_id"],
                    "action": "data_import_update",
                    "changed_by": actor_ctx.user_id,
                    "changed_by_role": actor_ctx.role or "",
                    # The canonical reversible shape, the same one `update_student`
                    # writes, so one undo understands every path rather than each
                    # path inventing its own record of what it did.
                    "changes": {
                        field: {"previous": (item.get("previous") or {}).get(field),
                                "new": value}
                        for field, value in item["changes"].items()
                    },
                    "import_batch": batch_id,
                    "timestamp": actor_ctx.now_iso(),
                },
                school_id=actor_ctx.school_id,
                branch_id=actor_ctx.branch_id or "",
            )
            applied += 1
        except Exception as exc:
            failed += 1
            logger.warning("import update failed for %s: %s", item["admission_number"], exc)

    plan.update({
        "filename": record.get("filename", ""),
        "import_batch": batch_id,
        "applied": applied,
        "failed": failed,
        "message": (
            f"Filled {plan['fields_to_fill']} pieces of information across {applied} "
            f"students from '{record.get('filename')}'."
            + (f" {failed} could not be saved." if failed else "")
            + (
                " Ignored (outside your access): "
                + ", ".join(plan["columns_outside_your_access"]) + "."
                if plan.get("columns_outside_your_access") else ""
            )
        ),
    })
    return plan
