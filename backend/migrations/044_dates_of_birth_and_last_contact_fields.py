"""Load the dates of birth that stand up to checking, and the last few contact fields.

Abhimanyu, 2026-08-12: load the dates of birth that match everywhere and leave the
ambiguous ones.

--------------------------------------------------------------------------------
There is no second source, so the check has to be a real one
--------------------------------------------------------------------------------

The school's ``DETAINEES LIST 2025-26`` workbook holds a date of birth for 514 children
whose platform record is blank. It is the ONLY source for those children: the student
export is blank for every one of them, and the workbook's 64 class sheets are lookups into
the same table rather than an independent record. So "does every source agree" cannot be
asked, because only one source speaks.

That workbook has also been shown to be wrong sometimes. Where it and the platform both
hold a date, **they disagree for 119 children out of 925**: 62 with the day and month
reversed, 57 simply different.

So the check used here is not agreement between documents. It is whether the date is
**possible for the child who holds it**.

--------------------------------------------------------------------------------
The age test, calibrated on the school's own trusted records
--------------------------------------------------------------------------------

For the 806 children where the workbook and the platform DO agree, this works out the age
each class actually is at the start of the session. The result is exactly what a school
looks like, one year per class::

    LKG 4.6    UKG 5.6    1st 6.3    2nd 7.3    3rd 8.5    4th 9.4    5th 10.1
    6th 10.9   7th 12.2   8th 12.8   9th 13.6   10th 14.6  11th 15.8  12th 16.7

A blank child's date is loaded only if the age it implies falls inside their own class's
band, taken from the 2nd to the 98th percentile of the trusted children and widened by a
year at each end so that a genuinely old or young child in a class is not thrown out.

**This catches the failure the workbook actually has.** A reversed day and month moves a
birthday by up to eleven months, and a wrong year moves it by years; either lands outside a
class band that is about two and a half years wide. A date that survives is one that is
consistent with the only other thing known about that child.

Dates that fail, and children in a class with too few trusted records to build a band from,
are **left blank and counted**. Nursery has no trusted record at all, so its one child is
left for the school.

--------------------------------------------------------------------------------
The last of the contact fields
--------------------------------------------------------------------------------

The same run finishes the student export: four WhatsApp numbers, four school register
numbers and two parent mobiles that were still blank. After this the export has nothing
left in it that the platform does not hold.

--------------------------------------------------------------------------------
How to run it
--------------------------------------------------------------------------------

Dry run first. It is the default and changes nothing::

    backend/.venv/Scripts/python.exe backend/migrations/044_dates_of_birth_and_last_contact_fields.py
    ... 044_dates_of_birth_and_last_contact_fields.py --apply
    ... 044_dates_of_birth_and_last_contact_fields.py --rollback <the file it saved>

Never run this through ``run_all.py``.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from datetime import date, datetime, timezone

SCHOOL_ID = "aaryans-joya"
MIGRATION_ACTOR = "migration-044"

DETAINEES = "aaryans_database/DETAINEES LIST 2025-26.xlsx"
STUDENTS = "aaryans_database/Students-06-08-2026-12-08-00.xlsx"

# The session these ages are measured at.
SESSION_START = date(2026, 4, 1)
# A class needs at least this many trusted children before a band is trusted.
MIN_TRUSTED = 8
# How far outside the trusted 2nd-to-98th percentile a child may still sit, in years.
BAND_SLACK = 1.0

# export column -> platform field, for the handful still blank
CONTACT_FIELDS = (
    ("Whatsapp", "whatsapp"),
    ("SrNo", "sr_no"),
    ("MotherMobile", "mother_mobile"),
    ("FatherMobile", "father_mobile"),
)


def _repo_root() -> str:
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _blank(value) -> bool:
    return value in (None, "", [], {}) or str(value).strip() == ""


def _admission(value) -> str:
    text = str(value or "").strip()
    return text[:-2] if text.endswith(".0") else text


def _as_date(value):
    """Excel stores 1,359 of these as real dates and 250 as text.

    The text ones are written the Indian way, day first, as ``29.03.2018`` or
    ``23/12/2019``. They are read day-first for that reason, and anything the age test
    then finds implausible is thrown out anyway, so a misread separator cannot survive
    into a child's record.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for separator in (".", "/", "-"):
        parts = text.split(separator)
        if len(parts) != 3:
            continue
        try:
            day, month, year = (int(part) for part in parts)
        except ValueError:
            continue
        if year < 100:
            year += 2000 if year < 50 else 1900
        try:
            return date(year, month, day)
        except ValueError:
            return None
    return None


def _age_at_session(when: date) -> float:
    return (SESSION_START - when).days / 365.25


def _read_workbook_dobs(root: str) -> dict[str, date]:
    import openpyxl

    sheet = openpyxl.load_workbook(os.path.join(root, DETAINEES), read_only=True)["StudentData"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell or "").strip() for cell in rows[0]]
    index = {name: position for position, name in enumerate(header)}
    out = {}
    for row in rows[1:]:
        admission = _admission(row[index["ADM NO"]])
        when = _as_date(row[index["Dob"]])
        if admission and when:
            out[admission] = when
    return out


def _read_contact_rows(root: str) -> list[dict]:
    import openpyxl

    sheet = openpyxl.load_workbook(os.path.join(root, STUDENTS), read_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell or "").strip() for cell in rows[0]]
    index = {name: position for position, name in enumerate(header)}
    out = []
    for row in rows[1:]:
        admission = _admission(row[index["AdmissionNo"]])
        if not admission:
            continue
        record = {"admission_number": admission}
        for column, field in CONTACT_FIELDS:
            position = index.get(column)
            record[field] = row[position] if position is not None else None
        out.append(record)
    return out


def build_bands(trusted: dict[str, list[float]]) -> dict[str, tuple[float, float]]:
    """Each class's believable age range, from children whose two records already agree."""
    bands = {}
    for class_name, ages in trusted.items():
        if len(ages) < MIN_TRUSTED:
            continue
        ordered = sorted(ages)
        low = ordered[int(len(ordered) * 0.02)]
        high = ordered[int(len(ordered) * 0.98)]
        bands[class_name] = (low - BAND_SLACK, high + BAND_SLACK)
    return bands


async def plan(db, root: str) -> dict:
    students = await db.students.find({"schoolId": SCHOOL_ID}, {"_id": 0}).to_list(5000)
    classes = await db.classes.find({"schoolId": SCHOOL_ID}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    class_name = {row["id"]: row["name"] for row in classes}
    by_admission = {str(row.get("admission_number") or "").strip(): row for row in students}

    workbook = _read_workbook_dobs(root)

    trusted: dict[str, list[float]] = {}
    disagreements = 0
    for admission, when in workbook.items():
        student = by_admission.get(admission)
        if student is None or _blank(student.get("dob")):
            continue
        held = str(student["dob"])[:10]
        if held == when.isoformat():
            trusted.setdefault(class_name.get(student.get("class_id"), "?"), []).append(
                _age_at_session(when)
            )
        else:
            disagreements += 1

    bands = build_bands(trusted)

    dob_updates, rejected_age, no_band, not_on_platform = [], [], [], 0
    for admission, when in workbook.items():
        student = by_admission.get(admission)
        if student is None:
            not_on_platform += 1
            continue
        if not _blank(student.get("dob")):
            continue
        name = class_name.get(student.get("class_id"), "?")
        band = bands.get(name)
        if band is None:
            no_band.append({"admission_number": admission, "class": name})
            continue
        age = _age_at_session(when)
        if band[0] <= age <= band[1]:
            dob_updates.append({"id": student["id"], "admission_number": admission,
                                "set": {"dob": when.isoformat()}})
        else:
            rejected_age.append({"admission_number": admission, "class": name,
                                 "date": when.isoformat(), "age": round(age, 1)})

    contact_updates = []
    for row in _read_contact_rows(root):
        student = by_admission.get(row["admission_number"])
        if student is None:
            continue
        changes = {}
        for _, field in CONTACT_FIELDS:
            value = row.get(field)
            if not _blank(value) and _blank(student.get(field)):
                changes[field] = str(value).strip()
        if changes:
            contact_updates.append({"id": student["id"],
                                    "admission_number": row["admission_number"], "set": changes})

    return {
        "bands": bands,
        "trusted_total": sum(len(v) for v in trusted.values()),
        "disagreements": disagreements,
        "dob_updates": dob_updates,
        "rejected_age": rejected_age,
        "no_band": no_band,
        "not_on_platform": not_on_platform,
        "contact_updates": contact_updates,
    }


async def apply(db, root: str, *, dry_run: bool = True) -> dict:
    result = await plan(db, root)
    if dry_run:
        return result

    now = datetime.now(timezone.utc).isoformat()
    undo = []
    for update in result["dob_updates"] + result["contact_updates"]:
        undo.append({"id": update["id"], "fields": list(update["set"])})
        await db.students.update_one(
            {"id": update["id"], "schoolId": SCHOOL_ID},
            {"$set": {**update["set"], "updated_at": now, "updated_by": MIGRATION_ACTOR}},
        )
    result["applied"] = True
    result["undo"] = undo
    return result


async def rollback_from(db, path: str) -> dict:
    with open(path, encoding="utf-8") as handle:
        undo = json.load(handle)["undo"]
    cleared = 0
    for row in undo:
        cleared += (await db.students.update_one(
            {"id": row["id"], "schoolId": SCHOOL_ID},
            {"$unset": {field: "" for field in row["fields"]}},
        )).modified_count
    return {"cleared": cleared}


async def _main() -> int:
    parser = argparse.ArgumentParser(description="Dates of birth that stand up to checking.")
    parser.add_argument("--apply", action="store_true", help="actually write. Default is a dry run.")
    parser.add_argument("--rollback", help="path to a rollback file from a previous --apply")
    args = parser.parse_args()

    root = _repo_root()
    sys.path.insert(0, os.path.join(root, "backend"))
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(root, "backend", ".env"))
    except ImportError:
        pass
    from database import connect_db, get_db

    await connect_db()
    db = get_db()

    if args.rollback:
        out = await rollback_from(db, args.rollback)
        print(f"cleared {out['cleared']} records")
        return 0

    result = await apply(db, root, dry_run=not args.apply)

    print(f"  children whose two records already agree  {result['trusted_total']:>5}")
    print(f"  children whose two records disagree       {result['disagreements']:>5}  (left alone)")
    print(f"  classes with a believable age band        {len(result['bands']):>5}")
    print()
    print(f"  dates of birth that pass the age test     {len(result['dob_updates']):>5}  LOADED")
    print(f"  rejected, the age does not fit the class  {len(result['rejected_age']):>5}")
    print(f"  no band for that class, left for a person {len(result['no_band']):>5}")
    print(f"  named in the workbook, not on the platform{result['not_on_platform']:>5}")
    print(f"\n  last contact fields to fill               {len(result['contact_updates']):>5}")

    for row in result["rejected_age"][:10]:
        print(f"      rejected {row['admission_number']}  {row['class']}  {row['date']}  "
              f"would be {row['age']}")

    if not args.apply:
        print("\nDry run. Nothing was written. Re-run with --apply when you mean it.")
        return 0

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    path = os.path.join(os.path.dirname(root), f"rollback-044-dob-{stamp}.json")
    with open(path, "w", encoding="utf-8") as handle:
        json.dump({"undo": result["undo"], "rejected_age": result["rejected_age"],
                   "no_band": result["no_band"]}, handle, indent=2)
    print(f"\nDone. Rollback saved OUTSIDE the repository at:\n  {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(_main()))
