"""Fill the blank fields the school's own records can answer without a guess.

Abhimanyu, 2026-08-12: load everything we hold that carries no ambiguity, and leave the
rest for the school to fill in themselves.

Four blanks qualify. Each is filled ONLY where the platform currently holds nothing, so
this migration can never overwrite something a person typed.

--------------------------------------------------------------------------------
1. What each teacher teaches
--------------------------------------------------------------------------------

``staff.subject`` was blank for all 110 people. The platform already knows the answer:
its own ``subjects`` collection records who teaches each class's subject, and **45 members
of staff teach exactly one subject**. Not one of them teaches two, so there is nothing to
choose between.

This reads the platform's own data rather than ``more staff info.txt``, which is a typed
summary of photographs of a wall chart and names people by first name only. Three of the
first names in it belong to more than one member of staff.

--------------------------------------------------------------------------------
2. Where staff live
--------------------------------------------------------------------------------

The teacher export carries an address for 20 people and a town, county and postcode for
21. All 78 rows match exactly one staff record by name, with no ambiguity at all.

**Salary and qualification are NOT here, and cannot be.** The export has no salary column
of any kind, and exactly one of its 78 rows carries a qualification. There is no second
source. They go to the school to fill in.

--------------------------------------------------------------------------------
3. Whether a child is a boy or a girl
--------------------------------------------------------------------------------

201 children had no gender recorded. The school's ``DETAINEES LIST 2025-26`` workbook
records it for 111 of them, written as BOY and GIRL.

**Dates of birth from that same workbook are deliberately NOT loaded**, and this is worth
the sentence. The workbook holds a date of birth for 514 children whose platform record is
blank, which is tempting. But where the workbook and the platform BOTH hold a date, they
disagree for 119 children out of 925. 62 of those are the day and month the wrong way
round and 57 are simply different dates. A source that is wrong about one date in eight
cannot be trusted to fill in the ones nobody can check, and a wrong birthday follows a
child onto every certificate they are ever issued.

--------------------------------------------------------------------------------
4. The school's logo
--------------------------------------------------------------------------------

``school_settings.logo_url`` was empty. The school's logo already ships with the app at
``/aaryans-logo.jpg``; only the setting pointing at it was missing.

--------------------------------------------------------------------------------
How to run it
--------------------------------------------------------------------------------

Dry run first. It is the default and changes nothing::

    backend/.venv/Scripts/python.exe backend/migrations/042_fill_blanks_from_school_sources.py
    ... 042_fill_blanks_from_school_sources.py --apply
    ... 042_fill_blanks_from_school_sources.py --rollback <the file it saved>

Never run this through ``run_all.py``.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone

SCHOOL_ID = "aaryans-joya"
MIGRATION_ACTOR = "migration-042"

TEACHERS = "aaryans_database/Teachers-06-08-2026-12-09.xlsx"
DETAINEES = "aaryans_database/DETAINEES LIST 2025-26.xlsx"
LOGO_URL = "/aaryans-logo.jpg"

GENDER_WORDS = {"boy": "male", "girl": "female"}


def _repo_root() -> str:
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _norm(value) -> str:
    return " ".join(str(value or "").split()).casefold()


def _blank(value) -> bool:
    return value in (None, "", [], {}) or str(value).strip() == ""


def _read_teacher_rows(root: str) -> list[dict]:
    import openpyxl

    sheet = openpyxl.load_workbook(os.path.join(root, TEACHERS), read_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell or "").strip() for cell in rows[1]]
    index = {name: position for position, name in enumerate(header)}
    out = []
    for row in rows[2:]:
        if not any(row):
            continue
        def cell(name):
            position = index.get(name)
            return row[position] if position is not None else None
        full = " ".join(str(part).strip() for part in (cell("FirstName"), cell("LastName"))
                        if part not in (None, "")).strip()
        out.append({
            "name": full,
            "address": cell("Address"),
            "city": cell("City"),
            "state": cell("State"),
            "pincode": cell("Pincode"),
        })
    return out


def _read_detainee_genders(root: str) -> dict[str, str]:
    import openpyxl

    book = openpyxl.load_workbook(os.path.join(root, DETAINEES), read_only=True)
    sheet = book["StudentData"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell or "").strip() for cell in rows[0]]
    index = {name: position for position, name in enumerate(header)}
    out = {}
    for row in rows[1:]:
        admission = str(row[index["ADM NO"]] or "").strip()
        if admission.endswith(".0"):
            admission = admission[:-2]
        word = _norm(row[index["gender"]])
        if admission and word in GENDER_WORDS:
            out[admission] = GENDER_WORDS[word]
    return out


async def plan(db, root: str) -> dict:
    staff = await db.staff.find({"schoolId": SCHOOL_ID}, {"_id": 0}).to_list(5000)
    subjects = await db.subjects.find({"schoolId": SCHOOL_ID}, {"_id": 0}).to_list(5000)
    students = await db.students.find(
        {"schoolId": SCHOOL_ID}, {"_id": 0, "id": 1, "admission_number": 1, "gender": 1}
    ).to_list(5000)

    by_user_id = {row["user_id"]: row for row in staff if row.get("user_id")}
    by_name: dict[str, list] = defaultdict(list)
    for row in staff:
        by_name[_norm(row.get("name"))].append(row)

    # 1. subject, only where a member of staff teaches exactly one
    taught = defaultdict(set)
    for subject in subjects:
        if subject.get("teacher_id") and subject.get("name"):
            taught[subject["teacher_id"]].add(str(subject["name"]).strip())
    subject_updates, teaches_several = [], 0
    for teacher_id, names in taught.items():
        if len(names) != 1:
            teaches_several += 1
            continue
        row = by_user_id.get(teacher_id)
        if row and _blank(row.get("subject")):
            subject_updates.append({"id": row["id"], "name": row.get("name"),
                                    "set": {"subject": next(iter(names))}})

    # 2. where staff live
    address_updates, unmatched_staff, ambiguous_staff = [], 0, 0
    for source in _read_teacher_rows(root):
        matches = by_name.get(_norm(source["name"]), [])
        if not matches:
            unmatched_staff += 1
            continue
        if len(matches) > 1:
            ambiguous_staff += 1
            continue
        row = matches[0]
        changes = {}
        for field in ("address", "city", "state", "pincode"):
            value = source.get(field)
            if not _blank(value) and _blank(row.get(field)):
                changes[field] = str(value).strip()
        if changes:
            address_updates.append({"id": row["id"], "name": row.get("name"), "set": changes})

    # 3. boy or girl
    genders = _read_detainee_genders(root)
    gender_updates, gender_not_on_platform = [], 0
    by_admission = {str(row.get("admission_number") or "").strip(): row for row in students}
    for admission, value in genders.items():
        row = by_admission.get(admission)
        if row is None:
            gender_not_on_platform += 1
            continue
        if _blank(row.get("gender")):
            gender_updates.append({"id": row["id"], "admission_number": admission,
                                   "set": {"gender": value}})

    settings = await db.school_settings.find_one({"schoolId": SCHOOL_ID}, {"_id": 0})
    logo_needed = bool(settings) and _blank((settings or {}).get("logo_url"))

    return {
        "subject_updates": subject_updates,
        "teaches_several": teaches_several,
        "address_updates": address_updates,
        "unmatched_staff": unmatched_staff,
        "ambiguous_staff": ambiguous_staff,
        "gender_updates": gender_updates,
        "gender_not_on_platform": gender_not_on_platform,
        "logo_needed": logo_needed,
        "logo_before": (settings or {}).get("logo_url", ""),
    }


async def apply(db, root: str, *, dry_run: bool = True) -> dict:
    result = await plan(db, root)
    if dry_run:
        return result

    now = datetime.now(timezone.utc).isoformat()
    undo = {"staff": [], "students": [], "logo_before": result["logo_before"],
            "logo_applied": False}

    for update in result["subject_updates"] + result["address_updates"]:
        current = await db.staff.find_one({"id": update["id"], "schoolId": SCHOOL_ID}, {"_id": 0})
        undo["staff"].append({"id": update["id"],
                              "before": {key: (current or {}).get(key) for key in update["set"]}})
        await db.staff.update_one(
            {"id": update["id"], "schoolId": SCHOOL_ID},
            {"$set": {**update["set"], "updated_at": now, "updated_by": MIGRATION_ACTOR}},
        )

    for update in result["gender_updates"]:
        undo["students"].append({"id": update["id"], "before": {"gender": None}})
        await db.students.update_one(
            {"id": update["id"], "schoolId": SCHOOL_ID},
            {"$set": {**update["set"], "updated_at": now, "updated_by": MIGRATION_ACTOR}},
        )

    if result["logo_needed"]:
        await db.school_settings.update_one(
            {"schoolId": SCHOOL_ID},
            {"$set": {"logo_url": LOGO_URL, "updated_at": now}},
        )
        undo["logo_applied"] = True

    result["applied"] = True
    result["undo"] = undo
    return result


async def rollback_from(db, path: str) -> dict:
    with open(path, encoding="utf-8") as handle:
        undo = json.load(handle)["undo"]
    staff = students = 0
    for row in undo["staff"]:
        unset = {key: "" for key, value in row["before"].items() if value in (None, "")}
        keep = {key: value for key, value in row["before"].items() if value not in (None, "")}
        change = {}
        if unset:
            change["$unset"] = unset
        if keep:
            change["$set"] = keep
        if change:
            staff += (await db.staff.update_one({"id": row["id"], "schoolId": SCHOOL_ID},
                                                change)).modified_count
    for row in undo["students"]:
        students += (await db.students.update_one(
            {"id": row["id"], "schoolId": SCHOOL_ID}, {"$unset": {"gender": ""}}
        )).modified_count
    if undo.get("logo_applied"):
        await db.school_settings.update_one(
            {"schoolId": SCHOOL_ID}, {"$set": {"logo_url": undo.get("logo_before", "")}}
        )
    return {"staff": staff, "students": students}


async def _main() -> int:
    parser = argparse.ArgumentParser(description="Fill blanks the school's own records answer.")
    parser.add_argument("--apply", action="store_true", help="actually write. Default is a dry run.")
    parser.add_argument("--rollback", help="path to a rollback file from a previous --apply")
    args = parser.parse_args()

    root = _repo_root()
    sys.path.insert(0, os.path.join(root, "backend"))
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(root, "backend", ".env"))
    except ImportError:
        pass
    from database import connect_db, get_db

    await connect_db()
    db = get_db()

    if args.rollback:
        out = await rollback_from(db, args.rollback)
        print(f"put back {out['staff']} staff fields and {out['students']} student genders")
        return 0

    result = await apply(db, root, dry_run=not args.apply)

    print(f"  staff gaining the subject they teach   {len(result['subject_updates']):>5}")
    print(f"  staff gaining an address or a town     {len(result['address_updates']):>5}")
    print(f"  children gaining boy or girl           {len(result['gender_updates']):>5}")
    print(f"  school logo to set                     {'yes' if result['logo_needed'] else 'already set':>5}")
    print(f"\n  teacher rows matching no staff record  {result['unmatched_staff']:>5}")
    print(f"  teacher rows matching more than one    {result['ambiguous_staff']:>5}")
    print(f"  children named in the workbook, not on the platform {result['gender_not_on_platform']:>5}")
    print("\n  NOT loaded, deliberately: salary and qualification (in no file we hold), and")
    print("  514 dates of birth from the detainees workbook, which disagrees with the")
    print("  platform on 119 of the 925 children where both hold a date.")

    if not args.apply:
        print("\nDry run. Nothing was written. Re-run with --apply when you mean it.")
        return 0

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    path = os.path.join(os.path.dirname(root), f"rollback-042-blanks-{stamp}.json")
    with open(path, "w", encoding="utf-8") as handle:
        json.dump({"undo": result["undo"]}, handle, indent=2)
    print(f"\nDone. Rollback saved OUTSIDE the repository at:\n  {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(_main()))
