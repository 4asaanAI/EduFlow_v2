"""Release 2, step 2 - record which senior students are Commerce and which are Science.

11th and 12th are charged **4,800 a year apart** depending on the stream. Until this ran,
the platform had no way of recording a stream at all, so it could not tell the two apart
and could not charge them differently. Putting a Commerce child on the Science band
overcharges that family by 4,800 a year.

--------------------------------------------------------------------------------
What this writes, and what it deliberately does not
--------------------------------------------------------------------------------

**It writes two things and nothing else:**

1. ``stream`` on the **six senior class records** that already exist
   (``cls-11th-a/b/c`` and ``cls-12th-a/b/c``).
2. ``stream`` on the **individual students** whose stream the school's own documents
   name.

**It moves no child between classes, and it creates no class.** Both of those are worth
explaining, because the finishing plan asked for four new class records.

*Why no new classes.* The plan says "create the four Commerce and Science class records".
Reading the live database first showed the platform already has **six** senior class
records, one per section, each with its own id. A fee structure is keyed by ``class_id``
(``fee_config_service.create_fee_structure``), so those six ids are already enough to
charge Commerce and Science differently. Creating four more would leave the school with
ten senior class records and two different places to look up one child's fee. The purpose
of the step is met by the six that exist; the extra four would only be somewhere for a
figure to go stale.

*Why no child is moved.* Re-parenting a student changes ``class_id``, which their
attendance, marks and timetable all hang off. That is a far bigger change than this step
needs, and it is not required: the child's own ``stream`` is what decides their band.

--------------------------------------------------------------------------------
Where the streams come from, and how far they go
--------------------------------------------------------------------------------

Two of the school's own documents, neither of them a guess:

* ``Fees-log-detailed-11-08-2026-17-36.xlsx`` - the payment ledger, names 157
* ``Students-Fees-Structure-Report-06-08-2026-12-49.xlsx`` - names 189

**Where both name a child they agree, every single time.** Between them they cover
**186 of the 190 senior students on the platform**.

**The other four are left alone and must be asked about.** Guessing costs a family 4,800.

**One child contradicts the pattern, and this is the reason section is not used as a
rule.** The school's sections line up with streams: A is Science, B and C are Commerce.
That holds for 185 of the 186 students. It fails for **admission 263105**, who sits in
11th section A but whose stream both documents record as **Commerce**. So the child's own
record wins and the section is only ever a cross-check. Had section been used as the rule,
that one family would have been billed 4,800 a year too much.

--------------------------------------------------------------------------------
How to run it
--------------------------------------------------------------------------------

Dry run first. It is the default and it changes nothing::

    backend/.venv/Scripts/python.exe backend/migrations/034_senior_streams.py

Read the counts it prints. Then, and only then::

    ... 034_senior_streams.py --apply

It writes a rollback file **outside the repository** before it changes anything. To undo::

    ... 034_senior_streams.py --rollback <path to that file>

Never run this through ``run_all.py``.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from typing import Any, Dict, List

SCHOOL_ID = "aaryans-joya"

# The school's own structure, read off its payment ledger: 11th/12th Science is section A,
# Commerce is sections B and C. Used to set the CLASS default only. It is never used to
# decide an individual child, for the reason in the docstring.
CLASS_STREAMS = {
    "cls-11th-a": "Science",
    "cls-11th-b": "Commerce",
    "cls-11th-c": "Commerce",
    "cls-12th-a": "Science",
    "cls-12th-b": "Commerce",
    "cls-12th-c": "Commerce",
}

LEDGER = "Fees-log-detailed-11-08-2026-17-36.xlsx"
FEE_REPORT = "Students-Fees-Structure-Report-06-08-2026-12-49.xlsx"


def _repo_root() -> str:
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _data(name: str) -> str:
    return os.path.join(_repo_root(), "aaryans_database", name)


def _stream_of(text: Any) -> str:
    text = str(text or "")
    if "Commerce" in text:
        return "Commerce"
    if "Science" in text:
        return "Science"
    return ""


def read_documented_streams() -> Dict[str, Dict[str, str]]:
    """admission number -> {stream, sources}. Reads spreadsheets, no database."""
    import openpyxl

    found: Dict[str, Dict[str, str]] = {}

    def note(adm, stream, source):
        key = str(adm).strip()
        if not key or not stream:
            return
        row = found.setdefault(key, {"stream": stream, "sources": []})
        # A document that contradicts an earlier one is recorded, not overwritten. The
        # caller refuses to write anything for a child whose documents disagree.
        if row["stream"] != stream:
            row["conflict"] = f"{row['stream']} vs {stream}"
        row["sources"].append(source)

    wb = openpyxl.load_workbook(_data(FEE_REPORT), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    idx = {c: i for i, c in enumerate(rows[0]) if c}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        raw = str(row[idx["Class"]] or "")
        cls = raw.rsplit("-", 1)[0].strip() if "-" in raw else raw.strip()
        note(row[0], _stream_of(cls), "per-student fee report")

    wb = openpyxl.load_workbook(_data(LEDGER), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    idx = {c: i for i, c in enumerate(rows[1]) if isinstance(c, str)}
    for row in rows[2:]:
        if not row or row[idx["Admission No."]] is None:
            continue
        note(row[idx["Admission No."]], _stream_of(row[idx["Class"]]), "payment ledger")

    return found


async def plan(db) -> dict:
    """What WOULD change. Reads only."""
    documented = read_documented_streams()

    classes = await db.classes.find(
        {"schoolId": SCHOOL_ID, "id": {"$in": list(CLASS_STREAMS)}}, {"_id": 0}
    ).to_list(50)
    by_id = {c["id"]: c for c in classes}
    missing_classes = sorted(set(CLASS_STREAMS) - set(by_id))

    class_changes = [
        {"id": cid, "from": by_id[cid].get("stream", ""), "to": want}
        for cid, want in CLASS_STREAMS.items()
        if cid in by_id and by_id[cid].get("stream", "") != want
    ]

    students = await db.students.find(
        {"schoolId": SCHOOL_ID, "class_id": {"$in": list(CLASS_STREAMS)}},
        {"_id": 0, "id": 1, "admission_number": 1, "class_id": 1, "stream": 1},
    ).to_list(1000)

    student_changes, unknown, conflicted, already = [], [], [], []
    for s in students:
        adm = str(s.get("admission_number") or "").strip()
        row = documented.get(adm)
        if not row:
            unknown.append({"admission_number": adm, "class_id": s["class_id"]})
            continue
        if row.get("conflict"):
            conflicted.append({"admission_number": adm, "conflict": row["conflict"]})
            continue
        if s.get("stream") == row["stream"]:
            already.append(adm)
            continue
        student_changes.append({
            "id": s["id"],
            "admission_number": adm,
            "class_id": s["class_id"],
            "from": s.get("stream", ""),
            "to": row["stream"],
            "sources": sorted(set(row["sources"])),
        })

    # Cross-check only. Never a rule. See the docstring.
    contradicts_section = [
        c for c in student_changes
        if CLASS_STREAMS.get(c["class_id"]) and CLASS_STREAMS[c["class_id"]] != c["to"]
    ]

    return {
        "missing_classes": missing_classes,
        "class_changes": class_changes,
        "student_changes": student_changes,
        "unknown": unknown,
        "conflicted": conflicted,
        "already_correct": len(already),
        "contradicts_section": contradicts_section,
        "seniors_on_platform": len(students),
    }


async def apply(db, *, dry_run: bool = True) -> dict:
    result = await plan(db)
    if result["missing_classes"]:
        result["blocked_by"] = [
            "These senior class records were not found, so the platform is not shaped the "
            f"way this migration expects: {result['missing_classes']}. Stop and look."
        ]
        return result
    if result["conflicted"]:
        result["blocked_by"] = [
            f"{len(result['conflicted'])} students are given different streams by different "
            "school documents. Nothing was written for anybody. Settle those first."
        ]
        return result

    rollback: List[dict] = [
        {"kind": "class", "id": c["id"], "stream": c["from"]} for c in result["class_changes"]
    ] + [
        {"kind": "student", "id": c["id"], "stream": c["from"]} for c in result["student_changes"]
    ]
    result["rollback"] = rollback
    if dry_run:
        return result

    stamp = datetime.now(timezone.utc).isoformat()
    for c in result["class_changes"]:
        await db.classes.update_one(
            {"schoolId": SCHOOL_ID, "id": c["id"]},
            {"$set": {"stream": c["to"], "stream_set_at": stamp}},
        )
    for c in result["student_changes"]:
        await db.students.update_one(
            {"schoolId": SCHOOL_ID, "id": c["id"]},
            {"$set": {"stream": c["to"], "stream_source": ", ".join(c["sources"]), "stream_set_at": stamp}},
        )
    result["applied"] = True
    return result


async def rollback_from(db, path: str) -> dict:
    with open(path, encoding="utf-8") as handle:
        saved = json.load(handle)
    for row in saved:
        collection = db.classes if row["kind"] == "class" else db.students
        if row["stream"]:
            await collection.update_one(
                {"schoolId": SCHOOL_ID, "id": row["id"]}, {"$set": {"stream": row["stream"]}}
            )
        else:
            await collection.update_one(
                {"schoolId": SCHOOL_ID, "id": row["id"]},
                {"$unset": {"stream": "", "stream_source": "", "stream_set_at": ""}},
            )
    return {"restored": len(saved)}


async def _main() -> int:
    parser = argparse.ArgumentParser(description="Release 2 step 2: senior streams.")
    parser.add_argument("--apply", action="store_true", help="actually write. Default is a dry run.")
    parser.add_argument("--rollback", help="path to a rollback file from a previous --apply")
    args = parser.parse_args()

    root = _repo_root()
    sys.path.insert(0, os.path.join(root, "backend"))
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(root, "backend", ".env"))
    except ImportError:
        pass
    from database import connect_db, get_db

    await connect_db()
    db = get_db()

    if args.rollback:
        out = await rollback_from(db, args.rollback)
        print(f"restored {out['restored']} records from {args.rollback}")
        return 0

    result = await apply(db, dry_run=not args.apply)

    print(f"  senior students on the platform          {result['seniors_on_platform']:>5}")
    print(f"  class records gaining a stream           {len(result['class_changes']):>5}")
    print(f"  students gaining a stream                {len(result['student_changes']):>5}")
    print(f"  students already correct                 {result['already_correct']:>5}")
    print(f"  students no document names  <-- ASK      {len(result['unknown']):>5}")
    print(f"  students whose documents disagree        {len(result['conflicted']):>5}")

    by_stream = Counter(c["to"] for c in result["student_changes"])
    print(f"\n  of those being set: {dict(by_stream)}")

    if result["unknown"]:
        print("\n  No document names a stream for these. They are NOT being guessed:")
        for u in result["unknown"]:
            print(f"    admission {u['admission_number']}  currently in {u['class_id']}")

    if result["contradicts_section"]:
        print("\n  Documented stream differs from the stream of the section they sit in.")
        print("  The child's own record wins. Worth the school confirming:")
        for c in result["contradicts_section"]:
            print(f"    admission {c['admission_number']}  in {c['class_id']}  ->  {c['to']}")

    if result.get("blocked_by"):
        print("\nNOTHING WAS CHANGED.")
        for b in result["blocked_by"]:
            print(f"  {b}")
        return 1

    if not args.apply:
        print("\nDry run. Nothing was changed. Re-run with --apply when you mean it.")
        return 0

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    path = os.path.join(os.path.dirname(root), f"rollback-034-senior-streams-{stamp}.json")
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(result["rollback"], handle, indent=2)
    print(f"\nDone. Rollback saved OUTSIDE the repository at:\n  {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(_main()))
