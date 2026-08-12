"""The exports: who may download what, in which format, and how completely.

Written for UI Sweep Epic 10, Story 10.4, which added Excel alongside CSV and
changed packaging only - so the important tests then were the ones asserting
NOTHING ELSE MOVED, because a format option that quietly widened who can download
the school's data would be far worse than the inconvenience it fixed.

**Release 3, 2026-08-12, did move the gates, deliberately.** They were hand-written
role checks that predated the Release 2 permission table and disagreed with it in
both directions. They now derive from `services/profile_matrix.py`, so the tests
below assert the TABLE's answer rather than a second copy of it. Three changed from
403 to 200 and each says why in its own docstring. The wall that matters - the
management head and money - is asserted directly and did not move.

Release 3 also closed a live silent-truncation defect in every export. See the
last two tests in this file.
"""
from __future__ import annotations

import io

import pytest

from middleware.auth import create_jwt

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _bearer(payload: dict) -> dict:
    return {"Authorization": f"Bearer {create_jwt(payload)}"}


def _owner():
    return _bearer({"user_id": "x-owner", "role": "owner", "name": "Owner"})


def _principal():
    return _bearer({"user_id": "x-prin", "role": "admin", "sub_category": "principal",
                    "branch_id": "branch-a", "name": "Principal"})


def _accountant():
    return _bearer({"user_id": "x-acct", "role": "admin", "sub_category": "accountant",
                    "name": "Accountant"})


def _management():
    """Lalit Thomas. Holds the student and staff records and no money screen at all,
    which is the Release 2 wall an export must not get around."""
    return _bearer({"user_id": "x-mgmt", "role": "admin", "sub_category": "management",
                    "name": "Management"})


def _receptionist():
    """A DORMANT profile. It holds `student-database` in the permission table and
    still must not be able to download the whole roll."""
    return _bearer({"user_id": "x-recep", "role": "admin", "sub_category": "receptionist",
                    "name": "Front Desk"})


def _teacher():
    return _bearer({"user_id": "x-teach", "role": "teacher", "branch_id": "branch-a", "name": "Teacher"})


def _student():
    return _bearer({"user_id": "x-stu", "role": "student", "name": "Student"})


_TOUCHED = ("students", "staff", "fee_transactions", "student_attendance",
            "expenses", "enquiries", "classes")


@pytest.fixture(autouse=True)
def _clean(fake_db):
    saved = {n: list(getattr(fake_db, n).docs) for n in _TOUCHED}
    for n in _TOUCHED:
        getattr(fake_db, n).docs[:] = []
    yield
    for n in _TOUCHED:
        getattr(fake_db, n).docs[:] = saved[n]


def _seed_students(fake_db):
    fake_db.students.docs.extend([
        {"id": "s1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
         "name": "Asha Kumari", "admission_number": "25001", "is_active": True},
        {"id": "s2", "schoolId": "aaryans-joya", "branch_id": "branch-a",
         "name": "Bipin Sharma", "admission_number": "25002", "is_active": True},
    ])


# ── The new format works ────────────────────────────────────────────────────────

def test_students_as_xlsx_is_a_real_workbook_with_the_data(client, fake_db):
    _seed_students(fake_db)
    resp = client.get("/api/export/students?format=xlsx", headers=_owner())

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MIME)
    assert ".xlsx" in resp.headers["content-disposition"]

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    values = [[c.value for c in row] for row in wb.active.iter_rows()]
    flat = [str(v) for row in values for v in row if v is not None]
    assert "Asha Kumari" in flat
    assert "Name" in flat


def test_csv_is_still_the_default(client, fake_db):
    """Every existing caller - buttons, scripts, bookmarks - keeps working."""
    _seed_students(fake_db)
    resp = client.get("/api/export/students", headers=_owner())

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert "Asha Kumari" in resp.text


def test_an_unknown_format_falls_back_to_csv_rather_than_erroring(client, fake_db):
    """Matches how Epic 3 handled an unrecognised sort field: the client's option
    list is a convenience, never the enforcement."""
    _seed_students(fake_db)
    resp = client.get("/api/export/students?format=banana", headers=_owner())

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")


@pytest.mark.parametrize("path,headers_fn", [
    ("/api/export/students", _owner),
    ("/api/export/staff", _owner),
    ("/api/export/fee-transactions", _accountant),
    ("/api/export/expenses", _accountant),
    ("/api/export/enquiries", _principal),
    ("/api/export/attendance", _owner),
    ("/api/export/exam-results", _owner),
])
def test_every_export_offers_xlsx(client, path, headers_fn):
    resp = client.get(f"{path}?format=xlsx", headers=headers_fn())
    assert resp.status_code == 200, path
    assert resp.headers["content-type"].startswith(XLSX_MIME), path


def test_an_empty_export_still_produces_an_openable_workbook(client, fake_db):
    """A workbook with only headers must still open. A zero-byte download is the
    'failure that looks like success' defect in a new place."""
    resp = client.get("/api/export/students?format=xlsx", headers=_owner())
    assert resp.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.active is not None


# ── Who may download what ───────────────────────────────────────────────────────
#
# This heading read "Nothing else moved: the gates are exactly as they were" until
# 2026-08-12, which stopped being true when Release 3 pointed the gates at the
# permission table. The three lines that hold regardless, and are the real point:
# a student and an unauthenticated caller are refused everything, in every format,
# and no format option is ever a way around a permission.

@pytest.mark.parametrize("path", [
    "/api/export/students", "/api/export/staff", "/api/export/fee-transactions",
    "/api/export/expenses", "/api/export/enquiries", "/api/export/attendance",
    "/api/export/exam-results",
])
@pytest.mark.parametrize("fmt", ["", "?format=xlsx"])
def test_a_student_is_refused_every_export_in_every_format(client, path, fmt):
    """The format option must not become a way around an export permission."""
    resp = client.get(f"{path}{fmt}", headers=_student())
    assert resp.status_code == 403, f"{path}{fmt}"


@pytest.mark.parametrize("path", [
    "/api/export/students", "/api/export/staff", "/api/export/fee-transactions",
    "/api/export/expenses", "/api/export/enquiries",
])
def test_unauthenticated_is_refused_every_export(client, path):
    assert client.get(f"{path}?format=xlsx").status_code == 401


def test_a_teacher_still_cannot_export_students_or_staff(client, fake_db):
    """These were owner-or-principal before this story and must stay so."""
    _seed_students(fake_db)
    assert client.get("/api/export/students?format=xlsx", headers=_teacher()).status_code == 403
    assert client.get("/api/export/staff?format=xlsx", headers=_teacher()).status_code == 403


def test_the_principal_can_export_fees_and_expenses(client):
    """CHANGED on 2026-08-12 (Release 3). This asserted 403 and now asserts 200.

    It was pinning a hand-written owner-or-accountant gate that predated the
    Release 2 permission table and disagreed with it. Three settled decisions all
    say the principal holds this:

      * The table gives the principal EVERY screen and the finance domain,
        including fee-collection, financial-reports and expense-tracker
        (`services/profile_matrix.py`). The only things withheld from him are the
        three the registry marks owner-only: the branch records, the school
        settings and the year-end promotion.
      * On 2026-08-08 the fee-reminder screens were widened to the principal for
        exactly this reason - the narrower gate "made them look broken rather than
        forbidden".
      * Release 3's whole-school download is for the owner AND the principal and
        contains fees and payments, so Abhimanyu has already settled that the
        principal may take fee figures out of the platform.

    Sonu and Lalit are the ones the wall runs between, and that wall is unchanged:
    `test_the_management_head_cannot_export_money` below is the test that matters.
    """
    assert client.get("/api/export/fee-transactions?format=xlsx", headers=_principal()).status_code == 200
    assert client.get("/api/export/expenses?format=xlsx", headers=_principal()).status_code == 200


def test_the_management_head_cannot_export_money(client):
    """The Release 2 wall, restated as a download.

    Lalit holds no finance screen, so he holds no finance export. This is the test
    that would catch an export becoming a way around the permission table, which is
    the thing Abhimanyu asked for by name.
    """
    assert client.get("/api/export/fee-transactions?format=xlsx", headers=_management()).status_code == 403
    assert client.get("/api/export/expenses?format=xlsx", headers=_management()).status_code == 403


def test_the_management_head_can_export_the_lists_he_works_in(client, fake_db):
    """The other half of the same decision. He runs the student and staff records
    daily and could not download either, because the old gate was owner-or-principal."""
    _seed_students(fake_db)
    assert client.get("/api/export/students?format=xlsx", headers=_management()).status_code == 200
    assert client.get("/api/export/staff?format=xlsx", headers=_management()).status_code == 200
    assert client.get("/api/export/enquiries?format=xlsx", headers=_management()).status_code == 200


def test_the_accountant_head_can_export_the_student_list(client, fake_db):
    """He holds the School Directory, a fee belongs to a child, and he reconciles
    the two. The old gate refused him the list he works in every day."""
    _seed_students(fake_db)
    assert client.get("/api/export/students?format=xlsx", headers=_accountant()).status_code == 200


def test_a_dormant_profile_is_refused_even_though_it_holds_the_screen(client, fake_db):
    """The front desk is granted `student-database` in the table but is DORMANT.

    Mirroring screen access alone would hand the front desk a download of all 1,876
    children the day that profile is switched on. Release 2 was explicit that it
    must not be what gives a dormant profile new powers, and a whole-roll download
    is a bigger thing than a paged screen, so switching a profile on has to include
    deciding what it may export.
    """
    _seed_students(fake_db)
    assert client.get("/api/export/students?format=xlsx", headers=_receptionist()).status_code == 403


def test_salary_is_still_withheld_from_the_staff_export(client, fake_db):
    """The staff query projects salary out. A new format must not reintroduce it."""
    fake_db.staff.docs.append({
        "id": "st1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
        "name": "Teacher One", "staff_type": "teacher", "is_active": True,
        "salary": 91234,
    })
    resp = client.get("/api/export/staff?format=xlsx", headers=_owner())
    assert resp.status_code == 200
    assert b"91234" not in resp.content


# ── An export is complete or it does not exist (Release 3) ────────────────────

def test_an_export_no_longer_stops_at_a_hidden_row_limit(client, fake_db):
    """The defect this closed, on 2026-08-12.

    Every export used to end in a hardcoded `to_list(N)` - students at 2,000 - with
    no count, no total and no warning. The roll is 1,876 children, so the next
    intake would have started dropping children out of the downloaded file silently.

    2,500 rows here is above the old ceiling and must all come back.
    """
    fake_db.students.docs.extend([
        {"id": f"s-{i}", "schoolId": "aaryans-joya", "branch_id": "branch-a",
         "name": f"Child {i}", "admission_number": f"A{i}", "is_active": True}
        for i in range(2500)
    ])

    resp = client.get("/api/export/students", headers=_owner())

    assert resp.status_code == 200
    body = resp.content.decode()
    # One header line plus a line per child. Trailing newline is not a row.
    assert len([ln for ln in body.splitlines() if ln.strip()]) == 2501
    # The child that the old 2,000-row ceiling would have dropped.
    assert "Child 2499" in body


def test_too_many_rows_refuses_the_download_instead_of_shortening_it(client, fake_db, monkeypatch):
    """If a list ever outgrows one file, the request FAILS and says so.

    A short file is the worst version of this bug: it leaves the building, gets
    filed as a record and reconciled against, and nothing on it says it is partial.
    """
    monkeypatch.setattr("routes.exports.EXPORT_MAX_ROWS", 10)
    fake_db.students.docs.extend([
        {"id": f"s-{i}", "schoolId": "aaryans-joya", "branch_id": "branch-a",
         "name": f"Child {i}", "is_active": True}
        for i in range(11)
    ])

    resp = client.get("/api/export/students", headers=_owner())

    assert resp.status_code == 413
    detail = resp.json()["detail"]
    # It has to say what to do next, and that no half-file was produced.
    assert "date range" in detail
    assert "no file at all" in detail


def test_the_excel_export_does_not_stop_at_the_document_builders_own_row_limit(
    client, fake_db, monkeypatch
):
    """The same silent-truncation fault, one layer further down (2026-08-12).

    `_read_all` was made complete-or-refused, but the Excel branch then handed the
    rows to the shared document builder, whose own ceiling of 5,000 rows TRUNCATES
    and appends a note near the bottom of the sheet. The payment ledger is about
    10,700 rows, so an Excel download of it was quietly losing more than half while
    the route above believed it had shipped everything.

    The builder's ceiling is dropped to 5 here so the wiring is what is under test,
    not openpyxl's speed. Every row must still be in the workbook, and the file must
    carry no "only the first N rows" note.
    """
    from openpyxl import load_workbook

    monkeypatch.setattr("services.document_builder.MAX_ROWS", 5)
    fake_db.students.docs.extend([
        {"id": f"s-{i}", "schoolId": "aaryans-joya", "branch_id": "branch-a",
         "name": f"Child {i}", "admission_number": f"A{i}", "is_active": True}
        for i in range(20)
    ])

    resp = client.get("/api/export/students?format=xlsx", headers=_owner())

    assert resp.status_code == 200
    sheet = load_workbook(io.BytesIO(resp.content)).active
    written = "\n".join(
        str(c.value) for row in sheet.iter_rows() for c in row if c.value is not None
    )
    assert "Child 19" in written
    assert "only the first" not in written


# ── The other 28 tables: packaging what a screen is already showing ───────────

def _table_body(**over):
    body = {"title": "Vendors", "headers": ["Name", "Trade"],
            "rows": [["Sharma Works", "Furniture"], ["Joya Stationers", "Books"]]}
    body.update(over)
    return body


def test_table_export_needs_a_login(client):
    resp = client.post("/api/export/table", json=_table_body())
    assert resp.status_code == 401


def test_table_export_returns_the_rows_the_screen_sent(client):
    """Most screens have no server export behind them, so the screen fetches every
    row it is showing through its own gated list endpoint and posts them here to be
    turned into a file. This route reads nothing and widens nothing."""
    resp = client.post("/api/export/table", json=_table_body(), headers=_owner())
    assert resp.status_code == 200
    body = resp.content.decode()
    assert "Name,Trade" in body
    assert "Joya Stationers" in body


def test_table_export_can_be_a_workbook(client):
    from openpyxl import load_workbook
    resp = client.post("/api/export/table", json=_table_body(format="xlsx"), headers=_owner())
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MIME
    sheet = load_workbook(io.BytesIO(resp.content)).active
    assert "Joya Stationers" in "\n".join(
        str(c.value) for row in sheet.iter_rows() for c in row if c.value is not None
    )


def test_table_export_refuses_too_many_rows_rather_than_shortening(client, monkeypatch):
    monkeypatch.setattr("routes.exports.TABLE_EXPORT_MAX_ROWS", 3)
    resp = client.post(
        "/api/export/table",
        json=_table_body(rows=[["a", "b"] for _ in range(4)]),
        headers=_owner(),
    )
    assert resp.status_code == 413
    assert "no file at all" in resp.json()["detail"]


def test_table_export_needs_column_headings(client):
    resp = client.post("/api/export/table", json=_table_body(headers=[]), headers=_owner())
    assert resp.status_code == 400


def test_a_cell_a_spreadsheet_would_run_is_shown_instead(client):
    """Excel treats a cell opening with = or @ as something to EXECUTE. These rows
    come from a screen, so this is the one export path where such a value can arrive
    from outside. Phone numbers and negative amounts are left alone on purpose - see
    the note beside `_FORMULA_STARTERS`."""
    resp = client.post(
        "/api/export/table",
        json=_table_body(rows=[["=HYPERLINK(\"http://x\")", "+919812345678"],
                               ["@SUM(A1)", "-4500"]]),
        headers=_owner(),
    )
    assert resp.status_code == 200
    body = resp.content.decode()
    assert "'=HYPERLINK" in body
    assert "'@SUM" in body
    # Untouched: prefixing these would turn real numbers into text.
    assert "+919812345678" in body and "'+91" not in body
    assert "-4500" in body and "'-4500" not in body
