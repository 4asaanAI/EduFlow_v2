"""The whole school in one Excel file, for Aman and Adesh only.

Release 3, item B. One download holding a separate sheet per area: children, staff,
fees and payments, attendance, exam results, classes, transport, expenses, enquiries.

WHAT THESE TESTS ARE ACTUALLY GUARDING. This is the single largest copy of the
school's data the platform can produce, so the two ways it can go wrong are the two
that matter in this release:

  1. A SHORT SHEET. Nine tabs is exactly the shape where one coming back empty or
     stopping early would go unnoticed - nobody scrolls to the bottom of tab five to
     check. So no sheet is ever trimmed: over the ceiling the whole request is refused
     and says so, and every sheet writes its own row count on its first line.
  2. THE WRONG HANDS. Owner and principal only, which is Abhimanyu's decision of
     2026-08-12. It is checked on the route, in Flo's tool, and again by the tool's
     leadership-only domain, and the tests below pin all three.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

from ai import tool_functions_v2
from middleware.auth import create_jwt
from routes import exports
from server import app

client = TestClient(app)

OWNER = {"id": "own-1", "role": "owner", "name": "Aman"}
PRINCIPAL = {"id": "pri-1", "role": "admin", "sub_category": "principal", "name": "Adesh"}
ACCOUNTANT = {"id": "acc-1", "role": "admin", "sub_category": "accountant", "name": "Sonu"}
MANAGEMENT = {"id": "mgt-1", "role": "admin", "sub_category": "management", "name": "Lalit"}
TEACHER = {"id": "tch-1", "role": "teacher", "name": "Teacher"}


def _bearer(user):
    return {"Authorization": f"Bearer {create_jwt(user)}"}


@pytest.fixture(autouse=True)
def _some_of_everything(fake_db):
    """A row or two in each area, so an empty sheet is distinguishable from a missing
    one. The counts differ on purpose: a bug that hands every sheet the same rows
    would pass if they all held three."""
    seed = {
        "students": [
            {"id": f"s-{i}", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "name": f"Child {i}", "admission_number": f"ADM{i}", "is_active": True,
             "class_id": "c-1", "route_zone_id": "r-1"}
            for i in range(7)
        ],
        "staff": [
            {"id": "t-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "name": "Mr Sharma", "staff_type": "teaching", "is_active": True},
        ],
        "fee_transactions": [
            {"id": "f-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "student_id": "s-1", "amount": 1200, "status": "paid"},
            {"id": "f-2", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "student_id": "s-2", "amount": 900, "status": "unpaid"},
        ],
        "student_attendance": [
            {"id": "a-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "student_id": "s-1", "date": "2026-08-01", "status": "present"},
        ],
        "exam_results": [
            {"id": "e-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "student_id": "s-1", "exam_id": "x-1", "marks_obtained": 80, "max_marks": 100},
        ],
        "classes": [
            {"id": "c-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "name": "5", "section": "A", "class_teacher_id": "t-1"},
        ],
        "transport_routes": [
            {"id": "r-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "route_name": "Joya town", "stops": ["Gate", "Chowk"], "is_active": True},
        ],
        "expenses": [
            {"id": "x-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "date": "2026-08-01", "category": "Power", "amount": 5000},
        ],
        "enquiries": [
            {"id": "q-1", "schoolId": "aaryans-joya", "branch_id": "branch-a",
             "student_name": "New Child", "status": "open", "created_at": "2026-08-01T00:00:00"},
        ],
    }
    saved = {}
    for name, docs in seed.items():
        col = getattr(fake_db, name)
        saved[name] = list(col.docs)
        col.docs[:] = docs
    yield
    for name, docs in saved.items():
        getattr(fake_db, name).docs[:] = docs


def _sheets(content: bytes) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content))
    return {ws.title: [[c.value for c in row] for row in ws.iter_rows()] for ws in wb.worksheets}


# ── One file, every area ─────────────────────────────────────────────────────

def test_the_owner_gets_one_file_with_a_sheet_for_every_area():
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    assert resp.status_code == 200

    sheets = _sheets(resp.content)
    assert list(sheets) == [name for _key, name in exports.WHOLE_SCHOOL_SHEETS]


def test_the_principal_gets_it_too():
    resp = client.get("/api/export/school-workbook", headers=_bearer(PRINCIPAL))
    assert resp.status_code == 200


def test_every_sheet_says_how_many_rows_it_holds():
    """A person cannot count 1,876 rows by eye, so the sheet has to tell them. Without
    this line an area that came back empty looks exactly like an area with no records
    in it."""
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    sheets = _sheets(resp.content)

    assert sheets["Children"][0][0] == "Children: 7 rows"
    assert sheets["Fees and Payments"][0][0] == "Fees and Payments: 2 rows"
    assert sheets["Transport"][0][0] == "Transport: 1 rows"


def test_the_sheets_hold_different_data_not_the_same_rows_nine_times():
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    sheets = _sheets(resp.content)

    # Row 3 is the heading row (row 1 is the count, row 2 is blank).
    assert sheets["Children"][2][0] == "Name"
    assert sheets["Transport"][2][0] == "Route"
    assert sheets["Classes"][2][0] == "Class"


def test_the_response_says_the_counts_without_opening_the_file():
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    counts = resp.headers.get("X-Export-Row-Counts", "")
    assert "Children: 7" in counts
    assert "Staff: 1" in counts


def test_a_class_carries_its_strength_and_its_teachers_name():
    """The class sheet is one of the two new builders, so its own reading is pinned.
    A class list without its strength is a list of names, not a record of the school."""
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    row = _sheets(resp.content)["Classes"][3]
    assert row[0] == "5" and row[1] == "A"
    assert row[2] == "Mr Sharma"
    assert row[3] == 7


def test_a_bus_route_carries_its_stops_and_its_riders():
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    row = _sheets(resp.content)["Transport"][3]
    assert row[0] == "Joya town"
    # The stops are a list on the record. Joined rather than dropped: a route sheet
    # without its stops is not a record of a route.
    assert row[3] == "Gate, Chowk"
    assert row[8] == 7


# ── Nothing is ever trimmed ──────────────────────────────────────────────────

def test_a_sheet_over_the_ceiling_refuses_rather_than_shipping_short(monkeypatch):
    """The whole request fails and says no file was produced. The alternative - a
    workbook whose fifth tab quietly stops - is this release's defining fault in the
    one place nobody would look for it."""
    monkeypatch.setattr(exports, "EXPORT_MAX_ROWS", 3)

    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))

    assert resp.status_code == 413
    detail = resp.json()["detail"]
    assert "students" in detail
    assert "no file at all" in detail.lower()


def test_the_workbook_builder_itself_also_refuses_rather_than_trimming():
    """Two refusals guard this, one behind the other. The reading refuses first (the
    test above). If a builder is ever added that reads some other way, the workbook
    still cannot be handed a sheet it would have to shorten."""
    from services.document_builder import DocumentBuildError, build_workbook

    with pytest.raises(DocumentBuildError) as exc:
        build_workbook(
            sheets=[{"name": "Children", "headers": ["Name"], "rows": [["a"], ["b"], ["c"]]}],
            max_rows=2,
        )
    assert "Children" in str(exc.value)
    assert "no file was produced" in str(exc.value).lower()


def test_money_comes_out_as_a_number_so_the_office_can_add_it_up():
    """"Rs 12,400" is text to a spreadsheet and will not add. The office downloads
    the fee sheet in order to total it, so the amount has to be a number."""
    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))
    fees = _sheets(resp.content)["Fees and Payments"]
    amounts = [row[4] for row in fees[3:]]
    assert amounts == [1200, 900]


def test_the_document_builders_own_5000_row_trim_does_not_apply(monkeypatch):
    """The shared builder trims at 5,000 rows and adds a note. That is right for a
    letter and wrong for a record, so the export's refusing ceiling is the only one in
    force. Dropped to 2 here, well under the 7 children, so the wiring is under test."""
    monkeypatch.setattr("services.document_builder.MAX_ROWS", 2)

    resp = client.get("/api/export/school-workbook", headers=_bearer(OWNER))

    assert resp.status_code == 200
    assert _sheets(resp.content)["Children"][0][0] == "Children: 7 rows"


# ── Owner and principal only ─────────────────────────────────────────────────

def test_workbook_unauthenticated_returns_401():
    assert client.get("/api/export/school-workbook").status_code == 401


@pytest.mark.parametrize("user", [ACCOUNTANT, MANAGEMENT, TEACHER],
                         ids=["accountant", "management", "teacher"])
def test_everyone_else_is_refused(user):
    """Sonu may export the ledger and Lalit may export the children, and neither may
    have the file that holds BOTH plus everything else. Abhimanyu, 2026-08-12."""
    resp = client.get("/api/export/school-workbook", headers=_bearer(user))
    assert resp.status_code == 403


def test_endpoint_wrong_role_returns_403():
    headers = _bearer({"id": "u1", "user_id": "u1", "role": "student", "name": "T"})
    assert client.get("/api/export/school-workbook", headers=headers).status_code == 403


# ── The same rules through Flo ───────────────────────────────────────────────

@pytest.fixture
def _fake_s3(monkeypatch):
    monkeypatch.setenv("S3_BUCKET", "eduflow-test-bucket")
    from services import document_export

    class _Stored:
        bucket, key, etag, sha256 = "b", "aaryans-joya/uploads/x/x.xlsx", "e", "s"

    monkeypatch.setattr(document_export, "upload_bytes", lambda **kw: _Stored())


async def test_flo_makes_the_same_file_and_reads_back_every_sheets_count(_fake_s3):
    result = await tool_functions_v2.tool_export_whole_school_workbook({}, OWNER)

    assert result["success"] is True
    assert result["data"]["row_counts"]["Children"] == 7
    assert result["data"]["sheets"] == len(exports.WHOLE_SCHOOL_SHEETS)
    # The counts are IN the sentence, not only in the data. The reply is what a person
    # reads; a file block that says nothing about its size is how a short one is filed.
    assert "Children 7" in result["message"]


@pytest.mark.parametrize("user", [ACCOUNTANT, MANAGEMENT], ids=["accountant", "management"])
async def test_flo_refuses_everyone_but_the_two_and_offers_the_smaller_thing(_fake_s3, user):
    result = await tool_functions_v2.tool_export_whole_school_workbook({}, user)

    assert result["success"] is False
    # `denied`, so the reply reads as a permission answer and not a fault.
    assert result["denied"] is True
    assert "owner" in result["message"].lower()


def test_the_tool_is_leadership_only_in_the_registry():
    """Belt and braces, on purpose. The function checks the role itself AND the tool
    sits in the leadership-only domain, so both would have to fail together for this
    file to reach the wrong desk."""
    assert "export_whole_school_workbook" in tool_functions_v2.LEADERSHIP_ONLY_TOOL_NAMES
    entry = tool_functions_v2.TOOL_REGISTRY["export_whole_school_workbook"]
    assert entry["access_domain"] == "leadership"
    # Read-class, and it must never grow a confirm window: an export needs no approval
    # (Abhimanyu, 2026-08-12).
    assert entry["dispatch_type"] == "read"
    assert not entry.get("requires_confirmation")
