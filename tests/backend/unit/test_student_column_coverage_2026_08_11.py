"""Every column of the school's own student export has a decision, and every import
lands on the field the record already uses.

Abhimanyu, 2026-08-11: carry the previous system's extra student columns into EduFlow,
but only where they HOLD data or matter even while empty, like blood group, because the
school intends to fill exactly those gaps through this platform. Do not carry across the
previous vendor's empty filler.

**The defect this file exists to stop.** Before today, eleven entries in
`STUDENT_FIELD_MAP` pointed at invented target names for facts the records already hold:
`whatsapp_phone` beside the real `whatsapp`, `father_phone` beside `father_mobile`,
`aadhaar_number` beside `aadhaar_no`. Importing the school's own export would have
written a SECOND copy of each, and every screen reads the original, so the import would
have appeared to work and changed nothing visible. That is the same shape as the
missing-data report that claimed 1,842 children had no date of birth while 1,055 had one
under another name.

So the rule pinned here is simple: **an import target must be a field the record already
knows about.** A name that is not in `student_service.UPDATABLE_FIELDS` is a new field
and needs a decision, not a typo.
"""

from __future__ import annotations

import re
from pathlib import Path

import pytest

from services.data_import_service import (
    ADMISSION_HEADERS,
    DELIBERATELY_NOT_IMPORTED,
    FINANCE_IMPORT_FIELDS,
    NON_FINANCE_IMPORT_FIELDS,
    PROTECTED_FIELDS,
    STUDENT_FIELD_MAP,
)
from services.student_service import UPDATABLE_FIELDS

EXPORT = (
    Path(__file__).resolve().parents[3]
    / "aaryans_database"
    / "Students-06-08-2026-12-08-00.xlsx"
)


def _norm(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


# ── The rule that stops a second copy of a fact ──────────────────────────────

def test_every_import_target_is_a_field_the_record_already_has():
    stray = sorted(set(STUDENT_FIELD_MAP.values()) - UPDATABLE_FIELDS)
    assert not stray, (
        "these import targets are not fields of a student record, so importing them "
        f"would create a second copy of something: {stray}"
    )


def test_no_header_is_both_mapped_and_refused():
    both = sorted(set(STUDENT_FIELD_MAP) & DELIBERATELY_NOT_IMPORTED)
    assert not both, f"mapped and refused at the same time: {both}"


@pytest.mark.parametrize(
    "header,expected",
    [
        # The eleven that were wrong. Each pair is one fact, and the right-hand name is
        # the one the 2026-08-06 load actually wrote onto 1,878 children.
        ("whatsapp", "whatsapp"),
        ("alternatenumber", "alternate_number"),
        ("fathermobile", "father_mobile"),
        ("mothermobile", "mother_mobile"),
        ("aadharno", "aadhaar_no"),
        ("penno", "pen_no"),
        ("registrationno", "registration_no"),
        ("bankaccountno", "bank_account_no"),
        ("previousschool", "attended_school"),
    ],
)
def test_the_renamed_targets_point_at_the_real_field(header, expected):
    assert STUDENT_FIELD_MAP[header] == expected


# ── Coverage of the school's actual file ─────────────────────────────────────

@pytest.mark.skipif(not EXPORT.exists(), reason="the school's export is not in this checkout")
def test_every_column_of_the_schools_export_has_a_decision():
    import openpyxl

    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    headers = list(next(wb.active.iter_rows(values_only=True)))
    wb.close()

    decided = set(STUDENT_FIELD_MAP) | set(DELIBERATELY_NOT_IMPORTED) | set(ADMISSION_HEADERS)
    undecided = [h for h in headers if _norm(h) not in decided]
    assert not undecided, (
        "these columns of the school's own export are neither imported nor explicitly "
        f"refused, so they would be reported as unrecognised: {undecided}"
    )


@pytest.mark.skipif(not EXPORT.exists(), reason="the school's export is not in this checkout")
def test_every_column_carrying_real_data_is_imported_or_refused_on_purpose():
    """A column with data that is merely unmapped is a silent gap. One that is refused
    is a decision. This asserts there is no third category."""
    import openpyxl

    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    it = wb.active.iter_rows(values_only=True)
    headers = list(next(it))
    rows = [r for r in it if r and r[0] is not None]
    wb.close()

    populated = [
        h for i, h in enumerate(headers)
        if any(i < len(r) and r[i] not in (None, "") for r in rows)
    ]
    mapped = set(STUDENT_FIELD_MAP) | set(ADMISSION_HEADERS)
    missing = [h for h in populated if _norm(h) not in mapped and _norm(h) not in DELIBERATELY_NOT_IMPORTED]
    assert not missing, f"columns holding real data with nowhere to go: {missing}"


# ── The things that must stay out, by name ───────────────────────────────────

@pytest.mark.parametrize("header", [
    "isrtestudent",       # decides whether a child owes ANY school fee
    "paidfees", "balancefees", "discount", "fine",
    "status", "dropout", "class", "section",
    "username", "sid",    # the previous system's identifiers, one of them a login name
    "scholarshippassword",
    "guardianname", "guardianmobile",
])
def test_these_are_never_imported(header):
    assert header not in STUDENT_FIELD_MAP
    assert header in DELIBERATELY_NOT_IMPORTED


def test_right_to_education_cannot_be_set_by_spreadsheet():
    """It is a government entitlement that removes a child's school fee entirely. The
    platform's own rule is that an import never touches what a family owes, and this is
    the flag that decides whether they owe anything at all. Editable one child at a
    time; never in bulk from a stale sheet."""
    assert "isrtestudent" in DELIBERATELY_NOT_IMPORTED
    assert "is_rte_student" not in set(STUDENT_FIELD_MAP.values())
    # ...but it must remain correctable on a single record, or a wrongly flagged child
    # cannot be fixed at all.
    assert "is_rte_student" in UPDATABLE_FIELDS


def test_money_and_enrolment_stay_protected():
    for field in ("fees", "balance", "paid_fees", "discount", "status", "admission_number"):
        assert field in PROTECTED_FIELDS


# ── The new fields the school will fill through the platform ─────────────────

@pytest.mark.parametrize("field", [
    "stream",          # 11th and 12th differ by 4,800 a year on this alone
    "blood_group",
    "height_cm", "weight_kg",
])
def test_empty_but_important_fields_are_carried_anyway(field):
    """Blank in every export so far. Kept because the school plans to fill them here,
    and a column that is blank today is not a column nobody wants."""
    assert field in UPDATABLE_FIELDS


def test_stream_is_importable_because_the_school_will_supply_it_in_bulk():
    assert STUDENT_FIELD_MAP["stream"] == "stream"


# ── Profile scoping survives the rename ──────────────────────────────────────

def test_the_bank_fields_are_still_the_accountants_alone():
    for field in ("bank_name", "bank_account_no", "bank_ifsc", "bank_branch", "account_holder"):
        assert field in FINANCE_IMPORT_FIELDS
        assert field not in NON_FINANCE_IMPORT_FIELDS, (
            f"{field} reached the management profile - the rename must not have widened it"
        )


def test_the_contact_numbers_reminders_are_sent_to_are_still_the_accountants():
    # He sends the fee reminders, so he keeps the numbers they go to.
    for field in ("phone", "whatsapp", "alternate_number", "father_mobile", "mother_mobile"):
        assert field in FINANCE_IMPORT_FIELDS


def test_management_gets_the_new_columns_but_not_the_bank_ones():
    assert "stream" in NON_FINANCE_IMPORT_FIELDS
    assert "blood_group" in NON_FINANCE_IMPORT_FIELDS
    assert "bank_ifsc" not in NON_FINANCE_IMPORT_FIELDS
