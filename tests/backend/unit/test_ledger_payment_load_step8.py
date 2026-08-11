"""Release 2 step 8: reading the school's payment ledger.

The two mistakes that would matter here are silent ones: reading past the summary row and
treating a different table as payments, and loading the same receipts twice so a family's
balance halves overnight.
"""

from __future__ import annotations

import importlib.util
import sys
from os.path import abspath, dirname, join

_ROOT = dirname(dirname(dirname(dirname(abspath(__file__)))))
sys.path.insert(0, join(_ROOT, "backend"))

_spec = importlib.util.spec_from_file_location(
    "migration_039", join(_ROOT, "backend", "migrations", "039_load_payments_from_ledger.py")
)
migration = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(migration)


def test_the_offices_wording_is_grouped_into_periods_the_platform_can_use():
    assert migration.classify("composite fees 1st q. (apr , may, jun) (apr)") == ("Composite Fee", "q1")
    assert migration.classify("composite fee 2 qtr (july, august, september)") == ("Composite Fee", "q2")
    assert migration.classify("composite fee 3 qtr (oct, nov, dec) (oct)") == ("Composite Fee", "q3")
    assert migration.classify("composite fee 4 qtr (jan, feb, march) 2 (jan)") == ("Composite Fee", "q4")


def test_every_transport_month_is_recognised_including_the_ones_billed_late():
    for short, full in [("apr", "april"), ("may", "may"), ("jul", "july"), ("aug", "august"),
                        ("sep", "september"), ("oct", "october"), ("nov", "november"),
                        ("dec", "december"), ("jan", "january"), ("feb", "february"),
                        ("mar", "march")]:
        assert migration.classify(f"transport fees {short}") == ("Transport Fee", f"transport-{full}")


def test_the_other_charges_the_school_makes():
    assert migration.classify("registration fee (apr)")[1] == "admission"
    assert migration.classify("admission fee (apr)")[1] == "admission"
    assert migration.classify("2025 - 2026 due fees")[1] == "previous-session"
    assert migration.classify("late fine")[0] == "Late Fine"
    assert migration.classify("chq bounse charges")[0] == "Cheque bounce charge"


def test_wording_nobody_planned_for_keeps_its_own_words_and_is_counted():
    # Never silently dropped and never forced into a period it does not belong to.
    head, period = migration.classify("SOMETHING THE OFFICE INVENTED")
    assert head == "SOMETHING THE OFFICE INVENTED"
    assert period == "other"


def test_the_ledgers_own_dates_are_read_as_day_first():
    # 07-08-2026 in this file is the 7th of August, not the 8th of July.
    assert migration._iso("07-08-2026") == "2026-08-07"
    assert migration._iso("23-01-2026") == "2026-01-23"
    assert migration._iso("") == ""
    assert migration._iso("not a date") == ""


def test_reading_stops_at_the_summary_row(tmp_path, monkeypatch):
    # The file has a SECOND table below the summary with different columns. Reading past
    # it would load a payment-mode breakdown as though every row were a receipt.
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Fees Log"] + [None] * 21)
    sheet.append(["Receipt No.", "Admission No."] + [None] * 20)
    real = [None] * 22
    real[0], real[1], real[7], real[9] = "R1", "1001", "late fine", 100
    sheet.append(real)
    total = [None] * 22
    total[0] = "Total"
    sheet.append(total)
    other_table = [None] * 22
    other_table[0], other_table[1] = "S.No.", "Amount"
    sheet.append(other_table)
    stray = [None] * 22
    stray[0], stray[1] = "1", "9750"
    sheet.append(stray)

    root = tmp_path
    (root / "aaryans_database").mkdir()
    book.save(str(root / migration.LEDGER))

    lines = migration.read_ledger(str(root))
    assert len(lines) == 1
    assert lines[0]["receipt"] == "R1"
    assert lines[0]["paid"] == 100


def test_a_ledger_with_no_summary_row_refuses_rather_than_reading_everything(tmp_path):
    import openpyxl
    import pytest

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Fees Log"] + [None] * 21)
    sheet.append(["Receipt No."] + [None] * 21)
    sheet.append(["R1"] + [None] * 21)
    (tmp_path / "aaryans_database").mkdir()
    book.save(str(tmp_path / migration.LEDGER))

    with pytest.raises(SystemExit) as exc:
        migration.read_ledger(str(tmp_path))
    assert "no 'Total' row" in str(exc.value)


def test_each_line_gets_a_key_that_survives_running_the_load_twice(tmp_path):
    # Two identical charges on one receipt are two real payments, so the key includes the
    # line's position. Without it the second would look like a duplicate and be skipped,
    # and the family would be short.
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Fees Log"] + [None] * 21)
    sheet.append(["Receipt No."] + [None] * 21)
    for _ in range(2):
        row = [None] * 22
        row[0], row[1], row[7], row[9] = "R9", "1001", "transport fees apr", 650
        sheet.append(row)
    total = [None] * 22
    total[0] = "Total"
    sheet.append(total)
    (tmp_path / "aaryans_database").mkdir()
    book.save(str(tmp_path / migration.LEDGER))

    keys = [line["ledger_key"] for line in migration.read_ledger(str(tmp_path))]
    assert len(keys) == 2
    assert len(set(keys)) == 2
