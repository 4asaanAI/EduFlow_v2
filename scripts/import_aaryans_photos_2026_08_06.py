"""Map the photographs out of the 6 Aug 2026 exports. DRY RUN BY DEFAULT.

    python scripts/import_aaryans_photos_2026_08_06.py            # report only
    python scripts/import_aaryans_photos_2026_08_06.py --apply    # writes

WHY THIS IS A SEPARATE SCRIPT. The first load (`import_aaryans_2026_08_06.py`) skipped
photographs, on the finding that the `Photo` column contained the literal word `View`
on every filled row. **That finding was wrong**, and Abhimanyu caught it by clicking
one: `View` is the link LABEL, and the cell's hyperlink TARGET is the real photograph.
Reading cell *values* — which is what `read_only=True` gives you — never sees it.

    openpyxl.load_workbook(path, data_only=True)     # <- hyperlinks ARE exposed
    openpyxl.load_workbook(path, read_only=True)     # <- cell.hyperlink is None

Three different shapes, all in the same pair of files:
  * student photo  — hyperlink target, absolute URL       (1,427)
  * teacher photo  — hyperlink target, absolute URL       (13)
  * parent photos  — RELATIVE path in the cell VALUE      (127 mother, 128 father, 1 guardian)
                     valid once prefixed with the CDN host; verified 200 image/jpeg

WHAT THIS WRITES: `photo_url` on students, and on matched staff. Nothing else.

WHAT IT DOES NOT WRITE, and why — see
`_bmad-output/planning-artifacts/data-load-anomalies-2026-08-06.md`:
  * PARENT photos. The `guardians` collection has no photo field. Rather than invent one
    silently (it would be invisible to the API, which serialises the Guardian model),
    the 256 paths are counted and recorded as anomaly C2.
  * Blank photos stay blank. 451 students have none; that is left as-is, per instruction.

A NOTE ON THE URLS, recorded as anomaly D2 and worth reading before relying on this:
every photograph lives on `cdn.vedmarg.com`, the school's PREVIOUS software vendor, and
opens with NO authentication. So (a) if that CDN goes away every photo on EduFlow breaks
at once, and (b) 1,427 children's photographs are publicly readable by anyone holding the
link. The right end state is re-hosting them in the school's own S3 under
`{school_id}/uploads/...` behind the authenticated route. Storing the URL is the useful
first step, not the finished job.

Confidentiality: prints counts only. No name, no URL, no image is written to console or
disk.
"""
from __future__ import annotations

import argparse
import datetime
import os
import uuid
from collections import Counter
from pathlib import Path

import certifi
import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / "aaryans_database"
load_dotenv(REPO / "backend" / ".env")

STUDENTS_XLSX = DATA / "Students-06-08-2026-12-08-00.xlsx"
TEACHERS_XLSX = DATA / "Teachers-06-08-2026-12-09.xlsx"
CDN_BASE = "https://cdn.vedmarg.com/"
SCHOOL_ID = os.environ.get("SCHOOL_ID", "aaryans-joya")


def norm_adm(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper() or None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def absolute(u):
    """Parent photos are stored relative; student/teacher photos already absolute."""
    if not u:
        return None
    u = str(u).strip()
    if u.startswith("http://") or u.startswith("https://"):
        return u
    return CDN_BASE + u.lstrip("/")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    print(f"=== {'APPLY (writing)' if args.apply else 'DRY RUN (nothing is written)'} ===\n")

    client = MongoClient(os.environ["MONGO_URL"], tlsCAFile=certifi.where(),
                         serverSelectionTimeoutMS=30000)
    db = client[os.environ["DB_NAME"]]

    # ---------- students ----------
    wb = openpyxl.load_workbook(STUDENTS_XLSX, data_only=True)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(hdr)}

    student_photos = {}     # admission_number -> url
    parent_photo_count = Counter()
    for row in ws.iter_rows(min_row=2):
        adm = norm_adm(row[ix["AdmissionNo"]].value)
        cell = row[ix["Photo"]]
        if adm and cell.hyperlink and cell.hyperlink.target:
            student_photos[adm] = absolute(cell.hyperlink.target)
        for col in ("MotherPhoto", "FatherPhoto", "GuardianPhoto"):
            if txt(row[ix[col]].value):
                parent_photo_count[col] += 1
    wb.close()

    live = {}
    for d in db.students.find({"schoolId": SCHOOL_ID}, {"_id": 0, "id": 1,
                                                        "admission_number": 1, "photo_url": 1}):
        a = norm_adm(d.get("admission_number"))
        if a:
            live[a] = d

    to_set, already, unmatched = [], 0, 0
    for adm, url in student_photos.items():
        doc = live.get(adm)
        if not doc:
            unmatched += 1
            continue
        if doc.get("photo_url") in (None, ""):
            to_set.append((doc["id"], url))
        elif doc.get("photo_url") == url:
            already += 1
        else:
            already += 1   # counted; never overwritten (rule 2 of the first load)

    print("STUDENT PHOTOGRAPHS")
    print(f"  hyperlinks in the spreadsheet     : {len(student_photos)}")
    print(f"  matched to a student, will be set : {len(to_set)}")
    print(f"  already had a photo (left alone)  : {already}")
    print(f"  no student for that admission no  : {unmatched}")
    print(f"  students with no photo in the file: {db.students.count_documents({'schoolId': SCHOOL_ID}) - len(student_photos)}"
          " (left blank, anomaly B3)")

    print("\nPARENT PHOTOGRAPHS — NOT written, no field exists (anomaly C2)")
    for k, v in parent_photo_count.items():
        print(f"  {k:16s} {v}")

    # ---------- teachers ----------
    wb = openpyxl.load_workbook(TEACHERS_XLSX, data_only=True)
    ws = wb.active
    thdr = [c.value for c in ws[2]]
    tix = {h: i for i, h in enumerate(thdr)}
    teacher_photos = []      # (name, phone10, url)
    for row in ws.iter_rows(min_row=3):
        cell = row[tix["Photo"]]
        if not (cell.hyperlink and cell.hyperlink.target):
            continue
        name = " ".join(filter(None, [txt(row[tix["FirstName"]].value),
                                      txt(row[tix["LastName"]].value)])).upper()
        phone = "".join(ch for ch in str(row[tix["Mobile"]].value or "") if ch.isdigit())[-10:]
        teacher_photos.append((name, phone, absolute(cell.hyperlink.target)))
    wb.close()

    staff = list(db.staff.find({"schoolId": SCHOOL_ID}, {"_id": 0, "id": 1, "name": 1,
                                                         "phone": 1, "photo_url": 1}))
    staff_set, staff_nomatch, staff_already = [], 0, 0
    for name, phone, url in teacher_photos:
        hit = None
        for s in staff:
            sp = "".join(ch for ch in str(s.get("phone") or "") if ch.isdigit())[-10:]
            # BOTH name and phone must agree — the same bar the teacher comparison used.
            if str(s.get("name") or "").strip().upper() == name and sp and sp == phone:
                hit = s
                break
        if not hit:
            staff_nomatch += 1
        elif hit.get("photo_url") in (None, ""):
            staff_set.append((hit["id"], url))
        else:
            staff_already += 1

    print("\nTEACHER PHOTOGRAPHS")
    print(f"  hyperlinks in the spreadsheet     : {len(teacher_photos)}")
    print(f"  matched on name AND phone, will set: {len(staff_set)}")
    print(f"  already had a photo (left alone)  : {staff_already}")
    print(f"  no confident match (NOT guessed)  : {staff_nomatch}")

    if not args.apply:
        print("\n--- DRY RUN. Nothing was written. Re-run with --apply. ---")
        client.close()
        return

    import json
    rb = Path(os.environ.get("ROLLBACK_DIR", Path.home())) / \
        f"aaryans_photos_rollback_{datetime.datetime.now():%Y%m%d_%H%M%S}.json"
    rb.write_text(json.dumps({
        "written_at": datetime.datetime.now().isoformat(),
        "undo_student_photo_ids": [i for i, _ in to_set],
        "undo_staff_photo_ids": [i for i, _ in staff_set],
    }, indent=1))
    print(f"\n  rollback manifest -> {rb}")

    ops = 0
    for sid_, url in to_set:
        db.students.update_one({"id": sid_, "schoolId": SCHOOL_ID},
                               {"$set": {"photo_url": url,
                                         "updated_at": datetime.datetime.now().isoformat()}})
        ops += 1
    print(f"  set photo_url on {ops} students")

    sops = 0
    for sid_, url in staff_set:
        db.staff.update_one({"id": sid_, "schoolId": SCHOOL_ID}, {"$set": {"photo_url": url}})
        sops += 1
    print(f"  set photo_url on {sops} staff")

    db.audit_logs.insert_one({
        "id": str(uuid.uuid4()), "schoolId": SCHOOL_ID, "branch_id": "branch-joya",
        "action": "bulk_import", "collection": "students",
        "entity_id": "aaryans-photos-2026-08-06",
        "changed_by": "layaa-ai-data-load", "changed_by_role": "system",
        "changes": {"student_photos_set": ops, "staff_photos_set": sops,
                    "parent_photos_not_stored": dict(parent_photo_count),
                    "source": "hyperlink targets, cdn.vedmarg.com"},
        "created_at": datetime.datetime.now().isoformat(),
    })
    print("  wrote 1 audit entry")
    client.close()


if __name__ == "__main__":
    main()
