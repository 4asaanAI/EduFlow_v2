"""
Add the new-joiner TEACHERS from the school's teacher export (owner corrections C4 + C5,
2026-08-06).

The owner's instruction, in their numbering:
  4. Of the 23 staff on EduFlow but missing from the export, the first 22 go on the
     "not in any spreadsheet" list; the LAST one, Yachika, is a NEW joiner whose record
     was never updated on the school's old system -> she belongs on the TEACHER list.
  5. Add the 12 teachers who are in the export but not on EduFlow -- they are new.
     Yachika appears in BOTH 4 and 5, so she must not be created twice.

The de-duplication in 4/5 is handled for free by working from the EXPORT and skipping
anyone who already matches a platform record: Yachika is in the export (as YACHIKA
MANGAT) and is created exactly once, from that single row.

RULES OBSERVED
  * never overwrite an existing staff record -- this script only CREATES
  * never delete -- the 22 who are absent from the export are untouched
  * dry-run by default; --apply writes and first writes a rollback manifest
  * counts in the transcript, names only to the gitignored aaryans_database/

Matching is on normalised NAME plus last-10-digits of MOBILE. The export's `StaffId`
column is entirely empty (anomaly G2), so there is no stable key to use instead; a row
is treated as "already on the platform" if EITHER the name or the phone matches, which
is deliberately cautious -- a false match means we skip (create nothing), never merge.

Usage:
    python scripts/import_aaryans_teachers_2026_08_06.py            # dry run
    python scripts/import_aaryans_teachers_2026_08_06.py --apply
"""
from __future__ import annotations

import argparse
import asyncio
import collections
import json
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import certifi
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / ".env")

SOURCE = ROOT / "aaryans_database" / "Teachers-06-08-2026-12-09.xlsx"
SCHOOL_ID = os.environ.get("SCHOOL_ID", "aaryans-joya")
BRANCH_ID = "branch-joya"
SOURCE_TAG = "teachers-export-2026-08-06"

# The export's header sits on the SECOND row; row 1 is a "Total Teachers: 78" banner.
HEADER_ROW = 2


def nk(s) -> str:
    return re.sub(r"[^a-z]", "", (s or "").lower())


def phone10(s) -> str:
    d = re.sub(r"\D", "", str(s or ""))
    return d[-10:] if len(d) >= 10 else ""


def read_export() -> list[dict]:
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[HEADER_ROW - 1]]
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = []
    for r in rows[HEADER_ROW:]:
        if not any(c not in (None, "") for c in r):
            continue

        def g(col):
            i = idx.get(col)
            if i is None or i >= len(r) or r[i] is None:
                return ""
            return str(r[i]).strip()

        first, last = g("FirstName"), g("LastName")
        name = (first + " " + last).strip()
        if not name:
            continue
        out.append({
            "name": name, "mobile": g("Mobile"), "whatsapp": g("Whatsapp"),
            "email": g("Email"), "gender": g("Gender").lower() or None,
            "username": g("Username"), "assigned_class": g("AssignedClass"),
            "section": g("Section"), "is_class_teacher": g("IsClassTeacher").lower() in ("1", "true", "yes"),
            "designation": g("Designation"), "qualification": g("Qualification"),
            "address": g("Address"), "city": g("City"), "state": g("State"),
            "pincode": g("Pincode"), "joining_date": g("JoiningDate"),
            "status": g("Status"), "photo": g("Photo"),
        })
    wb.close()
    return out


async def main(apply: bool) -> int:
    if not SOURCE.exists():
        print(f"ERROR: source not found: {SOURCE}")
        return 1
    export = read_export()

    client = AsyncIOMotorClient(os.environ["MONGO_URL"], tlsCAFile=certifi.where(), retryWrites=True)
    db = client[os.environ["DB_NAME"]]
    try:
        staff = [s async for s in db.staff.find(
            {"schoolId": SCHOOL_ID}, {"_id": 0, "id": 1, "name": 1, "phone": 1, "employee_id": 1})]
        by_name = {nk(s.get("name")) for s in staff}
        by_phone = {phone10(s.get("phone")) for s in staff if phone10(s.get("phone"))}

        new_rows, already = [], 0
        for row in export:
            if nk(row["name"]) in by_name or (phone10(row["mobile"]) and phone10(row["mobile"]) in by_phone):
                already += 1
                continue
            new_rows.append(row)

        # keep the TCH-NNNN sequence going rather than inventing a new prefix
        nums = [int(m.group(1)) for s in staff
                if (m := re.fullmatch(r"TCH-(\d+)", s.get("employee_id") or ""))]
        next_num = (max(nums) + 1) if nums else 1

        print("=" * 68)
        print("TEACHER ADDITIONS — source: Teachers-06-08-2026-12-09.xlsx")
        print("=" * 68)
        print(f"teachers in the export        : {len(export)}")
        print(f"staff already on the platform : {len(staff)}")
        print(f"  export rows already matched : {already}")
        print(f"  NEW -> would create         : {len(new_rows)}")
        print(f"employee ids would run        : TCH-{next_num:04d} .. TCH-{next_num + len(new_rows) - 1:04d}")
        ct = sum(1 for r in new_rows if r["is_class_teacher"])
        print(f"  of the new: class teachers  : {ct}, subject teachers: {len(new_rows) - ct}")
        with_phone = sum(1 for r in new_rows if phone10(r["mobile"]))
        print(f"  with a usable mobile number : {with_phone} of {len(new_rows)}")

        # names go to the gitignored folder only, never to the transcript
        (ROOT / "aaryans_database" / "_teachers_to_create.txt").write_text(
            "\n".join(f"{r['name']}\t{r['mobile']}\t{'class_teacher' if r['is_class_teacher'] else 'subject_teacher'}"
                      f"\t{r['assigned_class']}-{r['section']}" for r in new_rows))
        print("  (names -> aaryans_database/_teachers_to_create.txt)")

        if not apply:
            print("\nDRY RUN — nothing written. Re-run with --apply to write.")
            return 0

        docs = []
        for i, r in enumerate(new_rows):
            docs.append({
                "id": str(uuid.uuid4()),
                "schoolId": SCHOOL_ID,
                "branch_id": BRANCH_ID,
                "name": r["name"],
                "staff_type": "teacher",
                "role": "teacher",
                "sub_category": "class_teacher" if r["is_class_teacher"] else "subject_teacher",
                "employee_id": f"TCH-{next_num + i:04d}",
                "designation": r["designation"] or "Teacher",
                "department": None,
                "subject": None,
                "phone": r["mobile"] or None,
                "email": r["email"] or None,
                "gender": r["gender"],
                "address": r["address"] or None,
                "qualification": r["qualification"] or None,
                "join_date": r["joining_date"] or None,
                "salary": None,
                "casual_leave_balance": 12,
                "medical_leave_balance": 10,
                "earned_leave_balance": 15,
                # No login is created: the export carries no credentials, and inventing
                # passwords for real people is not ours to do. Same choice as the 9
                # office staff loaded earlier today.
                "is_active": (r["status"] or "").lower() != "inactive",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "source": SOURCE_TAG,
                "source_username": r["username"] or None,
            })

        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        manifest = ROOT / "aaryans_database" / f"_rollback_teachers_{stamp}.json"
        manifest.write_text(json.dumps({
            "script": Path(__file__).name, "written_at": stamp,
            "note": "To roll back, delete staff documents with exactly these ids.",
            "created_ids": [d["id"] for d in docs],
            "source_tag": SOURCE_TAG,
        }, indent=1))
        print(f"\nrollback manifest -> {manifest.name}")

        if docs:
            await db.staff.insert_many(docs)
        total = await db.staff.count_documents({"schoolId": SCHOOL_ID})
        print(f"created                       : {len(docs)}")
        print(f"staff on the platform now     : {total}")
        print("logins created                : 0 (no credentials in the export — deliberate)")
        return 0
    finally:
        client.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    sys.exit(asyncio.run(main(ap.parse_args().apply)))
