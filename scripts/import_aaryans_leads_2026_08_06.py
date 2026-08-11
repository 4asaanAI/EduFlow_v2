"""
Import the school's ENQUIRY / LEAD list (`aaryans_database/Leads-06-08-2026-16-55.xlsx`)
into the `enquiries` collection - owner request B, 2026-08-06.

These are admission-form leads: a child, the class applied for, both parents' names and
a contact number. All 102 rows are marked ACTIVE in the school's system, and the platform
had ZERO enquiries before this load, so nothing here can collide with existing work.

WHAT IS DELIBERATELY NOT DONE
  * No lead is matched to, or merged with, an enrolled student. 53 lead names also match
    the name of a child already on the roll, but a name is not an identity - that is the
    single rule this data transfer has been strictest about. They are loaded as enquiries
    and left for the school to link if they choose.
  * 5 of these leads share a name with the 30 applicants the owner said FAILED the
    admission test and were rejected. They are still loaded exactly as the school's own
    file has them (ACTIVE), because overriding a live school record on the strength of a
    name match would be the same mistake. The overlap is reported instead.

The Enquiry model carries only name/parent/phone/class. The file has much more - both
parents' names, qualifications, occupations, the previous school, address. Those are kept
on the document under clearly-named keys rather than thrown away; the enquiries routes
return raw documents, so they flow through the API without a model change.

Usage:
    python scripts/import_aaryans_leads_2026_08_06.py            # dry run
    python scripts/import_aaryans_leads_2026_08_06.py --apply
"""
from __future__ import annotations

import argparse
import asyncio
import collections
import json
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import certifi
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / ".env")

SOURCE = ROOT / "aaryans_database" / "Leads-06-08-2026-16-55.xlsx"
SCHOOL_ID = os.environ.get("SCHOOL_ID", "aaryans-joya")
BRANCH_ID = "branch-joya"
SOURCE_TAG = "leads-export-2026-08-06"

# The 30 applicants the owner confirmed were rejected after the admission test
# (correction C2). Used ONLY to report a name overlap, never to change a record.
#
# Their NAMES are not in this file: no child's name may enter the code repository, which
# is the reason `aaryans_database/` is gitignored and the reason every loader here prints
# counts rather than values. They are read from the confidential folder instead. If the
# file is absent the overlap check is skipped with a warning -- it is a reporting aid, so
# losing it must not stop 101 enquiries from loading.
REJECTED_FILE = ROOT / "aaryans_database" / "_rejected_applicants.json"


def load_rejected() -> list[str]:
    if not REJECTED_FILE.exists():
        print(f"WARNING: {REJECTED_FILE.name} not found - skipping the rejected-applicant "
              f"overlap check. The load itself is unaffected.")
        return []
    return json.loads(REJECTED_FILE.read_text()).get("names", [])


def nk(s) -> str:
    return re.sub(r"[^a-z]", "", (s or "").lower())


def s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).strip()


def parse_created(v) -> str | None:
    """'13-04-2026 04:00pm' -> ISO. Returns None if it cannot be read."""
    t = s(v)
    for fmt in ("%d-%m-%Y %I:%M%p", "%d-%m-%Y %I:%M:%S%p", "%d-%m-%Y"):
        try:
            return datetime.strptime(t, fmt).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    return None


def read_leads() -> list[dict]:
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [s(c) for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c not in (None, "") for c in r):
            continue
        d = dict(zip(hdr, list(r) + [None] * (len(hdr) - len(r))))
        name = s(d.get("Name"))
        if not name:
            continue
        out.append(d)
    wb.close()
    return out


async def main(apply: bool) -> int:
    if not SOURCE.exists():
        print(f"ERROR: source not found: {SOURCE}")
        return 1
    leads = read_leads()

    print("=" * 68)
    print("ENQUIRY / LEAD IMPORT - source: Leads-06-08-2026-16-55.xlsx")
    print("=" * 68)
    print(f"rows in the file          : {len(leads)}")
    applied = collections.Counter(s(d.get("Applied For")) or "(blank)" for d in leads)
    print(f"distinct classes applied  : {len(applied)}")
    print("  " + ", ".join(f"{k}={v}" for k, v in applied.most_common(8)) + " ...")
    with_phone = sum(1 for d in leads if re.sub(r"\D", "", s(d.get("Mobile No."))))
    print(f"with a contact number     : {with_phone}")
    print(f"with a date of birth      : {sum(1 for d in leads if s(d.get('DOB')))}")
    print(f"with a previous school    : {sum(1 for d in leads if s(d.get('Attended School')))}")

    rej = {nk(x) for x in load_rejected()}
    overlap = sorted({s(d.get("Name")) for d in leads if nk(d.get("Name")) in rej})
    print(f"\nnames shared with the 30 rejected applicants: {len(overlap)}")
    print("  loaded as the school's file has them (ACTIVE); reported, not overridden")

    client = AsyncIOMotorClient(os.environ["MONGO_URL"], tlsCAFile=certifi.where(), retryWrites=True)
    db = client[os.environ["DB_NAME"]]
    try:
        existing = {(nk(e.get("student_name")), re.sub(r"\D", "", e.get("phone") or ""))
                    async for e in db.enquiries.find({"schoolId": SCHOOL_ID},
                                                     {"_id": 0, "student_name": 1, "phone": 1})}
        print(f"\nenquiries already on the platform : {len(existing)}")

        to_write, skipped = [], 0
        seen_in_file = set()
        for d in leads:
            key = (nk(d.get("Name")), re.sub(r"\D", "", s(d.get("Mobile No."))))
            if key in existing or key in seen_in_file:
                skipped += 1
                continue
            seen_in_file.add(key)
            to_write.append(d)
        print(f"  already present / duplicate rows : {skipped}")
        print(f"  WOULD CREATE                     : {len(to_write)}")

        (ROOT / "aaryans_database" / "_leads_overlap_with_rejected.txt").write_text(
            "Leads whose NAME matches one of the 30 rejected applicants.\n"
            "Loaded as ACTIVE per the school's own file - please confirm which is right.\n\n"
            + "\n".join(overlap))
        print("  (overlap names -> aaryans_database/_leads_overlap_with_rejected.txt)")

        if not apply:
            print("\nDRY RUN - nothing written. Re-run with --apply to write.")
            return 0

        now = datetime.now(timezone.utc).isoformat()
        docs = []
        for d in to_write:
            father, mother = s(d.get("Father Name")), s(d.get("Mother Name"))
            docs.append({
                "id": (nid := str(uuid.uuid4())),
                "_id": nid,
                "schoolId": SCHOOL_ID,
                "branch_id": BRANCH_ID,
                "student_name": s(d.get("Name")),
                "parent_name": father or mother or None,
                "phone": re.sub(r"\D", "", s(d.get("Mobile No."))) or None,
                "class_applying": s(d.get("Applied For")),
                # every row in the file reads ACTIVE, i.e. a live enquiry nobody has
                # progressed yet - which is exactly what the platform calls "new"
                "status": "new",
                "source": "admission_form",
                "assigned_to": None,
                "created_at": parse_created(d.get("Creation Time")) or now,
                # --- detail the Enquiry model has no field for, kept rather than dropped
                "student_gender": s(d.get("Gender")).lower() or None,
                "student_dob": s(d.get("DOB")) or None,
                "father_name": father or None,
                "mother_name": mother or None,
                "father_mobile": re.sub(r"\D", "", s(d.get("Father Mobile"))) or None,
                "mother_mobile": re.sub(r"\D", "", s(d.get("Mother Mobile"))) or None,
                "father_occupation": s(d.get("Father Occupation")) or None,
                "mother_occupation": s(d.get("Mother Occupation")) or None,
                "father_qualification": s(d.get("Father Qualification")) or None,
                "mother_qualification": s(d.get("Mother Qualification")) or None,
                "address": s(d.get("Father Residential Address")) or s(d.get("Mother Residential Address")) or None,
                "city": s(d.get("City")) or None,
                "state": s(d.get("State")) or None,
                "pincode": s(d.get("Pincode")) or None,
                "previous_school": s(d.get("Attended School")) or None,
                "previous_class": s(d.get("Attended Class")) or None,
                "category": s(d.get("Category")) or None,
                "remark": s(d.get("Remark")) or None,
                "source_created_by": s(d.get("Created By")) or None,
                "source_status": s(d.get("Status")) or None,
                "source_tag": SOURCE_TAG,
            })

        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        manifest = ROOT / "aaryans_database" / f"_rollback_leads_{stamp}.json"
        manifest.write_text(json.dumps({
            "script": Path(__file__).name, "written_at": stamp,
            "note": "To roll back, delete enquiries documents with exactly these ids.",
            "source_tag": SOURCE_TAG, "created_ids": [d["id"] for d in docs],
        }, indent=1))
        print(f"\nrollback manifest -> {manifest.name}")

        if docs:
            await db.enquiries.insert_many(docs)
        total = await db.enquiries.count_documents({"schoolId": SCHOOL_ID})
        print(f"enquiries created          : {len(docs)}")
        print(f"enquiries on the platform  : {total}")
        return 0
    finally:
        client.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    sys.exit(asyncio.run(main(ap.parse_args().apply)))
