"""
Restore student HOUSE assignments for The Aaryans (owner-reported bug A1, 2026-08-06).

WHERE THE DATA CAME FROM
------------------------
`aaryans_database/DETAINEES LIST 2025-26.xlsx`. The file NAME is misleading: besides a
detainees sheet it carries a `StudentData` master sheet (1,743 rows) plus one sheet per
class, and every one of them has a `HOUSE` column holding the school's real house names
(AGAMYA / AGRIM / APRAJIT / ATULYA).

This matters because the obvious-looking source is WRONG in two ways:
  * the main student export's `HouseBlock` column is empty for all 1,878 rows, and
  * `backend/migrations/002_add_houses.py` creates four DIFFERENT houses
    (Shivaji/Tagore/Raman/Kalam) and assigns them ROUND-ROBIN by cursor order.
    Running it would have invented an assignment for every child. It must not be run.

The class sheets and the master sheet agree with each other on every single admission
number (0 conflicts across 1,362 students), which is why this load is considered safe.

RULES OBSERVED (see _bmad-output/HANDOFF-PROMPT-2026-08-06-evening.md)
  * match on ADMISSION NUMBER only, never on name
  * fill blanks only — never overwrite an existing house
  * never delete; students absent from the sheet are counted, not touched
  * dry-run by default; --apply writes, and always writes a rollback manifest first
  * print counts, never pupil names

Usage:
    python scripts/import_aaryans_houses_2026_08_06.py            # dry run
    python scripts/import_aaryans_houses_2026_08_06.py --apply    # writes
"""
from __future__ import annotations

import argparse
import asyncio
import collections
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import certifi
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / ".env")

SOURCE = ROOT / "aaryans_database" / "DETAINEES LIST 2025-26.xlsx"
MANIFEST_DIR = ROOT / "aaryans_database"

SCHOOL_ID = os.environ.get("SCHOOL_ID", "aaryans-joya")

# The school's four houses, exactly as the `houses` collection and the UI spell them.
# StudentDatabase.js drives its dropdown and colours off these strings, so the stored
# value must be the title-case NAME, not an id.
HOUSE_TITLE = {"AGAMYA": "Agamya", "AGRIM": "Agrim", "APRAJIT": "Aprajit", "ATULYA": "Atulya"}


def norm_adm(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "#N/A", "None"):
        return None
    if re.fullmatch(r"\d+(\.0)?", s):
        s = str(int(float(s)))
    return s if re.fullmatch(r"\d{4,8}", s) else None


def norm_house(v):
    """-> title-case house name, None for blank, or ('?', raw) for unrecognised."""
    if v is None:
        return None
    s = str(v).strip().upper().replace(".", "").replace(" ", "")
    if s in ("", "0", "0.0", "#N/A", "NONE"):
        return None
    return HOUSE_TITLE.get(s, ("?", str(v).strip()))


def read_sheet_houses():
    """Return (mapping adm->house, conflicts, unrecognised counter, sheets scanned)."""
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    found: dict[str, set] = collections.defaultdict(set)
    junk: collections.Counter = collections.Counter()
    sheets = 0
    for ws in wb.worksheets:
        hdr = adm_i = house_i = None
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
            cells = [(str(c).strip().upper() if c is not None else "") for c in row]
            ai = hi = None
            for ci, c in enumerate(cells):
                if c in ("ADM NO", "ADM.NO", "ADMNO", "ADM NO."):
                    ai = ci
                if c == "HOUSE":
                    hi = ci
            if ai is not None and hi is not None:
                hdr, adm_i, house_i = ri, ai, hi
                break
        if hdr is None:
            continue
        sheets += 1
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            adm = norm_adm(row[adm_i] if adm_i < len(row) else None)
            if not adm:
                continue
            h = norm_house(row[house_i] if house_i < len(row) else None)
            if h is None:
                continue
            if isinstance(h, tuple):
                junk[h[1]] += 1
                continue
            found[adm].add(h)
    wb.close()

    conflicts = {a: sorted(hs) for a, hs in found.items() if len(hs) > 1}
    clean = {a: next(iter(hs)) for a, hs in found.items() if len(hs) == 1}
    return clean, conflicts, junk, sheets


async def main(apply: bool) -> int:
    if not SOURCE.exists():
        print(f"ERROR: source not found: {SOURCE}")
        return 1

    clean, conflicts, junk, sheets = read_sheet_houses()
    print("=" * 68)
    print("HOUSE RESTORE — source: DETAINEES LIST 2025-26.xlsx")
    print("=" * 68)
    print(f"sheets carrying a HOUSE column      : {sheets}")
    print(f"admission numbers with a house      : {len(clean) + len(conflicts)}")
    print(f"  unambiguous                       : {len(clean)}")
    print(f"  CONFLICTING (skipped)             : {len(conflicts)}")
    if junk:
        print(f"  unrecognised house values         : {sum(junk.values())} -> {dict(junk)}")
    dist = collections.Counter(clean.values())
    print("  distribution: " + ", ".join(f"{h}={dist[h]}" for h in sorted(dist)))

    client = AsyncIOMotorClient(os.environ["MONGO_URL"], tlsCAFile=certifi.where(), retryWrites=True)
    db = client[os.environ["DB_NAME"]]
    try:
        # Guard: the four houses must already exist, and must be the school's real ones.
        house_names = {h.get("name") async for h in db.houses.find({"schoolId": SCHOOL_ID}, {"_id": 0, "name": 1})}
        missing = set(HOUSE_TITLE.values()) - house_names
        if missing:
            print(f"\nABORT: these houses do not exist in the database: {sorted(missing)}")
            return 1

        adms = list(clean.keys())
        to_write: list[tuple[str, str]] = []   # (student id, house)
        would_overwrite: list[tuple[str, str, str]] = []
        matched = 0
        for i in range(0, len(adms), 500):
            chunk = adms[i:i + 500]
            cursor = db.students.find(
                {"schoolId": SCHOOL_ID, "admission_number": {"$in": chunk}},
                {"_id": 0, "id": 1, "admission_number": 1, "house": 1},
            )
            async for st in cursor:
                matched += 1
                want = clean[st["admission_number"]]
                have = (st.get("house") or "").strip()
                if not have:
                    to_write.append((st["id"], want))
                elif have != want:
                    would_overwrite.append((st["admission_number"], have, want))

        total_students = await db.students.count_documents({"schoolId": SCHOOL_ID})
        print(f"\nstudents on the platform            : {total_students}")
        print(f"matched by admission number         : {matched}")
        print(f"  in the sheet but NOT on platform  : {len(adms) - matched}")
        print(f"  blank house -> WOULD WRITE        : {len(to_write)}")
        print(f"  already set and DIFFERENT (kept)  : {len(would_overwrite)}")
        print(f"students left with no house         : {total_students - len(to_write)}")

        if not apply:
            print("\nDRY RUN — nothing written. Re-run with --apply to write.")
            return 0

        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        manifest = MANIFEST_DIR / f"_rollback_houses_{stamp}.json"
        manifest.write_text(json.dumps({
            "script": Path(__file__).name,
            "written_at": stamp,
            "field": "house",
            "previous_value_for_all": None,
            "note": "To roll back, unset 'house' on exactly these student ids.",
            "student_ids": [sid for sid, _ in to_write],
            "conflicts_skipped": conflicts,
            "would_have_overwritten": would_overwrite,
        }, indent=1))
        print(f"\nrollback manifest -> {manifest.name}")

        written = 0
        for sid, house in to_write:
            res = await db.students.update_one(
                # re-assert the blank in the filter so a concurrent edit is never clobbered
                {"id": sid, "schoolId": SCHOOL_ID, "$or": [{"house": None}, {"house": ""}, {"house": {"$exists": False}}]},
                {"$set": {"house": house, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            written += res.modified_count
        print(f"students updated                    : {written}")

        verify = await db.students.count_documents({"schoolId": SCHOOL_ID, "house": {"$nin": [None, ""]}})
        print(f"verified: students now with a house : {verify}")
        by_house = await db.students.aggregate([
            {"$match": {"schoolId": SCHOOL_ID, "house": {"$nin": [None, ""]}}},
            {"$group": {"_id": "$house", "n": {"$sum": 1}}}, {"$sort": {"_id": 1}},
        ]).to_list(10)
        print("  " + ", ".join(f"{r['_id']}={r['n']}" for r in by_house))
        return 0
    finally:
        client.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write to the database (default is a dry run)")
    sys.exit(asyncio.run(main(ap.parse_args().apply)))
